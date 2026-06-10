#!/usr/bin/env python3
# Training Event Manager (plain Tkinter — no theme styling)
# -- early error capture --
import sys as _sys, os as _os, traceback as _tb_early
_script_dir = _os.path.dirname(_os.path.abspath(__file__))
_err_log = _os.path.join(_script_dir, "TEM_startup_error.log")
try:
    _sys.stderr = open(_err_log, "w", encoding="utf-8", buffering=1)
except Exception:
    pass

# Write a "started" marker immediately so we know Python reached this point
try:
    with open(_err_log, "w", encoding="utf-8") as _f:
        _f.write("Python started OK. Now importing modules...\n")
except Exception:
    pass

import os, re, time, itertools, json, tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import difflib
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional
import pandas as pd
from docx import Document

# Optional deps
try:
    import win32com.client as win32  # Outlook
except ImportError:
    win32 = None

try:
    from docx2pdf import convert as docx2pdf_convert  # needs Word installed
except ImportError:
    docx2pdf_convert = None

try:
    from tkcalendar import DateEntry  # calendar widget
except Exception:
    DateEntry = None

try:
    import winsound  # Windows sound
except Exception:
    winsound = None

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, HRFlowable)
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# For safe Excel appends (preserve workbook/tabs/formatting)
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
except Exception:
    load_workbook = None  # we'll warn if user tries to write XLSX without it

APP_TITLE = "Training Event Manager"
ORG_NAME = "University of Kentucky Center on Trauma and Children"
TEST_EMAIL = "jafish0@uky.edu"
TEST_NAME = "Joshua Fisherkeller"
MASTER_TRAINING_DEFAULT = r"\\UKHCDATA\dept\CTAC\TrainServ\Training Administration\Master Training List\Master Training List.xlsx"

DEFAULT_CERT_EMAIL = (
    "Hello <<FIRSTNAME>>! Thank you for attending “<<TRAINING>>”! Please find your attached certificate here.\n"
    "If there are any corrections needed or you have questions or concerns, please email me.\n\n"
    "Check out our upcoming events: https://ctac.uky.edu/events\n"
    "Resources & videos: https://www.youtube.com/@ukctac\n"
    "Facebook: https://www.facebook.com/profile.php?id=100039448224738\n\n"
    "Joshua Fisherkeller, MSW\n"
    "Education and Training Manager\n"
    "University of Kentucky Center on Trauma and Children\n"
    "College of Medicine, Department of Psychiatry\n"
    "3470 Blazer Parkway Suite 100\n"
    "Lexington, Kentucky 40509\n"
    "(859) 218-6941\n"
    "http://www.uky.edu/CTAC/"
)

PROGRAM_CHOICES = ["CTAC", "AWARE 3", "KY 6", "STS-ISC", "Stronger Connections", "RPC"]

DEFAULT_LC_EMAIL = (
    "Hello <<FIRSTNAME>>! Thank you for participating in \"<<TRAINING>>\"! "
    "Please find your attached certificate here.\n\n"
    "You attended <<HOURS_ATTENDED>> out of <<HOURS_TOTAL>> total hours.\n\n"
    "Your attendance record:\n"
    "<<ATTENDANCE_TABLE>>\n\n"
    "<<CEU_APPROVALS>>\n\n"
    "If there are any corrections needed or you have questions or concerns, please email me.\n\n"
    "Check out our upcoming events: https://ctac.uky.edu/events\n"
    "Resources & videos: https://www.youtube.com/@ukctac\n"
    "Facebook: https://www.facebook.com/profile.php?id=100039448224738\n\n"
    "Joshua Fisherkeller, MSW\n"
    "Education and Training Manager\n"
    "University of Kentucky Center on Trauma and Children\n"
    "College of Medicine, Department of Psychiatry\n"
    "3470 Blazer Parkway Suite 100\n"
    "Lexington, Kentucky 40509\n"
    "(859) 218-6941\n"
    "http://www.uky.edu/CTAC/"
)

# CEU approval texts
CEU_SWBOARD_TEXT   = "CTAC is approved by the Kentucky Board of Social Work as a sponsor for continuing education for Social Work (SW# KBSWSP 202536)."
CEU_PSYCHBOARD_TEXT = "The Center on Trauma and Children is approved by the Kentucky Board of Examiners of Psychology to offer continuing education for Psychologists."
CEU_LPCC_TEXT      = "This course is approved by the Board of Licensed Professional Counselors. CTAC maintains responsibility for this program and its content."
CEU_ASWB_TEMPLATE  = ("University of Kentucky Center on Trauma and Children, 112345, is approved as an ACE provider to offer social work continuing education "
                       "by the Association of Social Work Boards (ASWB) Approved Continuing Education (ACE) program. Regulatory boards are the final authority "
                       "on courses accepted for continuing education credit. ACE provider approval period: {period}. "
                       "Social workers completing this course receive {hours} continuing education credits.")
CEU_EILA_TEMPLATE  = "This course is approved for {hours} EILA Hours (EILA #: {number})."
CEU_FRSKY_TEMPLATE = "This course is approved for {hours} FRSKY Hours (FRSKY #: {number})."

KY_COUNTIES = [
    "Adair County","Allen County","Anderson County","Ballard County","Barren County","Bath County","Bell County",
    "Boone County","Bourbon County","Boyd County","Boyle County","Bracken County","Breathitt County","Breckinridge County",
    "Bullitt County","Butler County","Caldwell County","Calloway County","Campbell County","Carlisle County","Carroll County",
    "Carter County","Casey County","Christian County","Clark County","Clay County","Clinton County","Crittenden County",
    "Cumberland County","Daviess County","Edmonson County","Elliott County","Estill County","Fayette County","Fleming County",
    "Floyd County","Franklin County","Fulton County","Gallatin County","Garrard County","Grant County","Graves County",
    "Grayson County","Green County","Greenup County","Hancock County","Hardin County","Harlan County","Harrison County",
    "Hart County","Henderson County","Henry County","Hickman County","Hopkins County","Jackson County","Jefferson County",
    "Jessamine County","Johnson County","Kenton County","Knott County","Knox County","LaRue County","Laurel County",
    "Lawrence County","Lee County","Leslie County","Letcher County","Lewis County","Lincoln County","Livingston County",
    "Logan County","Lyon County","Madison County","Magoffin County","Marion County","Marshall County","Martin County",
    "Mason County","McCracken County","McCreary County","McLean County","Meade County","Menifee County","Mercer County",
    "Metcalfe County","Monroe County","Montgomery County","Morgan County","Muhlenberg County","Nelson County","Nicholas County",
    "Ohio County","Oldham County","Owen County","Owsley County","Pendleton County","Perry County","Pike County","Powell County",
    "Pulaski County","Robertson County","Rockcastle County","Rowan County","Russell County","Scott County","Shelby County",
    "Simpson County","Spencer County","Taylor County","Todd County","Trigg County","Trimble County","Union County",
    "Warren County","Washington County","Wayne County","Webster County","Whitley County","Wolfe County","Woodford County"
]

# ---------- helpers ----------

def smart_title(name: str) -> str:
    if not isinstance(name, str): return ""
    name = name.strip()
    if not name: return ""
    parts = []
    for part in name.split():
        hy = []
        for hp in part.split('-'):
            ap = [p.capitalize() if p else p for p in hp.split("'")]
            hp2 = "'".join(ap)
            hy.append(hp2.capitalize())
        parts.append("-".join(hy))
    s = " ".join(parts)
    s = re.sub(r"\bMc([a-z])", lambda m: "Mc" + m.group(1).upper(), s)
    s = re.sub(r"\bMac([a-z])", lambda m: "Mac" + m.group(1).upper(), s)
    return s

def extract_first(n: str) -> str:
    return n.strip().split()[0] if n and n.strip() else ""

def looks_like_email(s: str) -> bool:
    if not isinstance(s, str): return False
    s = s.strip()
    return bool(s) and ("@" in s) and (" " not in s) and (";" not in s) and ("," not in s)

def _smart_strip_df(df: pd.DataFrame) -> pd.DataFrame:
    return df.map(lambda x: x.strip() if isinstance(x, str) else x)

def md_to_html(text: str) -> str:
    text = (text.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"))
    text = text.replace("\r\n","\n").replace("\r","\n")
    return text.replace("\n","<br>")

def safe_filename(s: str) -> str:
    if s is None: return ""
    s = str(s)
    return re.sub(r'[\\/:*?"<>|]+', "_", s).strip()

def format_dates_short(dates_list):
    """Format ['08/27/2025','08/28/2025','08/29/2025'] -> '8/27/25, 8/28/25 and 8/29/25'"""
    def short(mdY):
        mdY = mdY.strip()
        try:
            dt = datetime.strptime(mdY, "%m/%d/%Y")
        except ValueError:
            try:
                dt = datetime.strptime(mdY, "%m/%d/%y")
            except ValueError:
                return mdY  # leave as-is if unknown
        return dt.strftime("%-m/%-d/%y") if os.name != "nt" else dt.strftime("%#m/%#d/%y")
    dates = [short(d) for d in dates_list if d.strip()]
    if not dates: return ""
    if len(dates) == 1: return dates[0]
    if len(dates) == 2: return f"{dates[0]} and {dates[1]}"
    return ", ".join(dates[:-1]) + f" and {dates[-1]}"

def count_psychologists(df: pd.DataFrame) -> int:
    if df is None or df.empty: return 0
    col = None
    for c in df.columns:
        if str(c).strip().lower() == "role":
            col = c; break
    if col is None:
        for c in df.columns:
            if str(c).strip().lower() in ("q7","q07","role (q7)","respondent role","participant role"):
                col = c; break
    if col is None:
        for c in df.columns:
            if "role" in str(c).strip().lower():
                col = c; break
    if col is None: return 0
    # Match: psych, psychology, psychologist, school psychology, school psychologist, etc.
    return int(df[col].astype(str).str.lower().str.contains(r"\bpsych", regex=True).sum())

# ---------- CSV parsing (Qualtrics-aware) ----------

def _detect_is_qualtrics_csv(path: str) -> bool:
    try:
        with open(path, "r", encoding="utf-8") as f:
            rows = list(itertools.islice(csv.reader(f), 0, 6))
        for r in rows:
            for cell in r:
                if isinstance(cell, str) and "importid" in cell.lower():
                    return True
    except Exception:
        pass
    try:
        sniff = pd.read_csv(path, dtype=str, nrows=6, header=None, keep_default_na=False)
        if sniff.map(lambda x: isinstance(x, str) and "importid" in x.lower()).any().any():
            return True
    except Exception:
        pass
    return False

def _parse_qualtrics_csv(path: str) -> pd.DataFrame:
    raw = pd.read_csv(path, dtype=str, header=None, keep_default_na=False)
    importid_idx = None
    for idx in range(min(6, len(raw))):
        row = raw.iloc[idx]
        if row.apply(lambda x: isinstance(x, str) and "importid" in x.lower()).any():
            importid_idx = idx; break
    if importid_idx is None:
        return pd.DataFrame(columns=["NAME","EMAIL","ROLE"])
    labels_idx = importid_idx - 1 if importid_idx >= 1 else 0
    tech_idx   = labels_idx - 1 if labels_idx >= 1 else None
    data_start = importid_idx + 1
    labels = list(raw.iloc[labels_idx])
    tech   = list(raw.iloc[tech_idx]) if tech_idx is not None else [None]*len(labels)

    # Find name/email columns
    name_idx = None; best = (9999,9999,9999)
    for i, lab in enumerate(labels):
        s = str(lab).strip(); sl = s.lower()
        if "name" in sl and "email" not in sl:
            recip = 1 if "recipient" in sl else 0
            is_exact = (sl == "name")
            tup = (recip, 0 if is_exact else 1, len(sl))
            if tup < best: best, name_idx = tup, i
    first_idx = next((i for i, lab in enumerate(labels) if isinstance(lab,str) and lab.strip().lower()=="recipient first name"), None)
    last_idx  = next((i for i, lab in enumerate(labels) if isinstance(lab,str) and lab.strip().lower()=="recipient last name"), None)

    email_idx = None; best_e = (9999,9999,9999,9999)
    for i, lab in enumerate(labels):
        s = str(lab).strip(); sl = s.lower()
        if "email" in sl:
            recip = 1 if "recipient" in sl else 0
            confirm = 1 if "confirm" in sl else 0
            techname = str(tech[i]).strip().lower() if tech[i] is not None else ""
            is_question = 0 if re.match(r"q\d+", techname) else 1
            tup = (recip, confirm, is_question, len(sl))
            if tup < best_e: best_e, email_idx = tup, i

    # Role column: prefer ImportId == Q7; fallback to label containing "role"
    role_idx = None
    if tech is not None:
        for i, t in enumerate(tech):
            if str(t).strip().lower() == "q7":
                role_idx = i; break
    if role_idx is None:
        for i, lab in enumerate(labels):
            if "role" in str(lab).strip().lower():
                role_idx = i; break

    if email_idx is None or (name_idx is None and not (first_idx is not None and last_idx is not None)):
        return pd.DataFrame(columns=["NAME","EMAIL","ROLE"])

    data = raw.iloc[data_start:, :].copy()
    if name_idx is not None:
        name_series = data.iloc[:, name_idx]
    else:
        first = data.iloc[:, first_idx].fillna("")
        last  = data.iloc[:, last_idx].fillna("")
        name_series = (first + " " + last).str.strip()
    email_series = data.iloc[:, email_idx]
    role_series = data.iloc[:, role_idx] if role_idx is not None else ""

    df = pd.DataFrame({"NAME": name_series, "EMAIL": email_series, "ROLE": role_series})
    df = _smart_strip_df(df)
    df["NAME"] = df["NAME"].apply(smart_title)
    df = df[(df["NAME"] != "") & (df["EMAIL"].apply(looks_like_email))]
    df = df.drop_duplicates(subset=["EMAIL"], keep="first")
    df = df.drop_duplicates(subset=["NAME"], keep="first").reset_index(drop=True)
    return df[["NAME","EMAIL","ROLE"]]

def _detect_columns_generic(df: pd.DataFrame):
    name_col = None; email_col = None; role_col = None
    for c in df.columns:
        lc = str(c).lower()
        if name_col is None and "name" in lc: name_col = c
        if email_col is None and "email" in lc: email_col = c
        if role_col is None and "role" in lc: role_col = c
    return name_col, email_col, role_col

def load_csv(path: str) -> pd.DataFrame:
    if _detect_is_qualtrics_csv(path):
        return _parse_qualtrics_csv(path)
    df = pd.read_csv(path, dtype=str, na_filter=False)
    df = _smart_strip_df(df)
    name_col, email_col, role_col = _detect_columns_generic(df)
    if not name_col:
        raise ValueError(f"Could not find a 'name' column in {os.path.basename(path)}")
    if email_col:
        df = df.rename(columns={email_col: "EMAIL"})
    else:
        df["EMAIL"] = ""
    df = df.rename(columns={name_col: "NAME"})
    if role_col:
        df = df.rename(columns={role_col: "ROLE"})
    else:
        df["ROLE"] = ""
    df["NAME"] = df["NAME"].apply(smart_title)
    df["EMAIL"] = df["EMAIL"].apply(lambda x: x.lower().strip() if isinstance(x, str) else "")
    df = df[(df["NAME"] != "") & df["EMAIL"].apply(looks_like_email)]
    df = df.drop_duplicates(subset=["EMAIL"], keep="first")
    df = df.drop_duplicates(subset=["NAME"], keep="first").reset_index(drop=True)
    return df[["NAME","EMAIL","ROLE"]]

# ---------- Outlook email ----------

def send_outlook_email(to, subject, body, attachment_path=None, html=True, cc=None):
    if not win32:
        raise RuntimeError("Outlook (pywin32) not available. Install with: pip install pywin32")
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = to
    if cc:
        mail.CC = cc
    mail.Subject = subject
    if html:
        mail.HTMLBody = body
    else:
        mail.Body = body
    if attachment_path and os.path.isfile(attachment_path):
        mail.Attachments.Add(attachment_path)
    mail.Send()

# ---------- DOCX replacement ----------

def _add_placeholder_variants(mapping_base: dict) -> dict:
    variants = {}
    for k, v in mapping_base.items():
        key = k.strip("<>").strip()
        for cand in (f"<<{key}>>", f"<< {key} >>", f"<<{key} >>", f"<< {key}>>"):
            variants[cand] = v
    return variants

def _replace_in_element(element, mapping: dict):
    paragraphs = element.xpath(".//w:p")
    for p in paragraphs:
        texts = p.xpath(".//w:t")
        if not texts: continue
        full_text = "".join(t.text or "" for t in texts)
        if not full_text: continue
        new_text = full_text
        for k, v in mapping.items():
            new_text = new_text.replace(k, str(v))
        if new_text != full_text:
            texts[0].text = new_text
            for t in texts[1:]: t.text = ""

def replace_placeholders_everywhere(doc: Document, mapping_base: dict):
    from docx.opc.constants import RELATIONSHIP_TYPE as RT
    mapping = _add_placeholder_variants(mapping_base)
    _replace_in_element(doc.element, mapping)
    for rel in doc.part.rels.values():
        try:
            if rel.reltype in (RT.HEADER, RT.FOOTER):
                _replace_in_element(rel.target_part.element, mapping)
        except Exception:
            pass

# ---------- Excel append (preserve workbook) ----------

def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = Font(name=src.font.name, size=src.font.size, bold=src.font.bold,
                        italic=src.font.italic, vertAlign=src.font.vertAlign,
                        underline=src.font.underline, strike=src.font.strike, color=src.font.color)
        dst.fill = PatternFill(fill_type=src.fill.fill_type, start_color=src.fill.start_color, end_color=src.fill.end_color)
        dst.border = Border(left=src.border.left, right=src.border.right, top=src.border.top, bottom=src.border.bottom,
                            diagonal=src.border.diagonal, diagonal_direction=src.border.diagonal_direction,
                            outline=src.border.outline, vertical=src.border.vertical, horizontal=src.border.horizontal)
        dst.alignment = Alignment(horizontal=src.alignment.horizontal, vertical=src.alignment.vertical,
                                  text_rotation=src.alignment.text_rotation, wrap_text=src.alignment.wrap_text,
                                  shrink_to_fit=src.alignment.shrink_to_fit, indent=src.alignment.indent)
        dst.number_format = src.number_format
        dst.protection = Protection(locked=src.protection.locked, hidden=src.protection.hidden)

def append_row_openpyxl(xlsx_path: str, sheet_name: str, row_values: list, hyperlink_col_idx: int = None, hyperlink_url: str = None):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to update .xlsx without changing other sheets.\nInstall with: pip install openpyxl")
    if not os.path.isfile(xlsx_path):
        raise RuntimeError("Selected Excel file does not exist or is unreachable.")
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]

    # append after last row
    last_row = ws.max_row
    target_row = last_row + 1
    style_src_row = last_row if last_row >= 1 else None

    for col_idx, value in enumerate(row_values, start=1):
        cell = ws.cell(row=target_row, column=col_idx, value=value)

        if style_src_row:
            src = ws.cell(row=style_src_row, column=col_idx)
            copy_cell_style(src, cell)

        # hyperlink handling
        if hyperlink_col_idx is not None and col_idx == hyperlink_col_idx and hyperlink_url:
            cell.hyperlink = hyperlink_url
            # try built-in hyperlink style if available; else make it blue + underline
            try:
                cell.style = "Hyperlink"
            except Exception:
                cell.font = Font(color="FF0563C1", underline="single")

    wb.save(xlsx_path)

# ---------- GUI helpers (autocomplete & scroll) ----------

class AutoCompleteCombobox(ttk.Combobox):
    """Simple type-ahead filter for values; still allows free typing."""
    def set_completion_list(self, completion_list):
        self._completion_list = sorted(list(set(completion_list)), key=str.lower)
        self["values"] = self._completion_list
        self.bind("<KeyRelease>", self._handle_keyrelease)

    def _handle_keyrelease(self, event):
        if event.keysym in ("BackSpace","Left","Right","Up","Down","Home","End","Tab","Return"):
            return
        typed = self.get().strip().lower()
        if not typed:
            self["values"] = self._completion_list
            return
        matches = [x for x in self._completion_list if typed in x.lower()]
        self["values"] = matches

class ScrollableFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.canvas = tk.Canvas(self, borderwidth=1, relief="sunken")
        self.vsb = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)
        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.frame = ttk.Frame(self.canvas)
        self.window_id = self.canvas.create_window((0,0), window=self.frame, anchor="nw")

        self.frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # Wheel bindings (Windows + Linux)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel_windows)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)

    def _on_frame_configure(self, _):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfigure(self.window_id, width=event.width)

    def _on_mousewheel_windows(self, event):
        self.canvas.yview_scroll(-int(event.delta/120), "units")

    def _on_mousewheel_linux(self, event):
        if event.num == 4: self.canvas.yview_scroll(-1, "units")
        elif event.num == 5: self.canvas.yview_scroll(1, "units")

# ---------- Main App ----------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1100x950")
        self.minsize(1000, 760)

        # State
        self.si_df = None
        self.so_df = None
        self.participant_count = tk.IntVar(value=0)
        self.event_dates = []  # list of strings mm/dd/yyyy

        # Standard tab: source type — "qualtrics_csv" (existing path) or
        # "roster_xlsx" (per-participant roster, Phase 3 of flexible ingest).
        self.std_source_type = tk.StringVar(value="qualtrics_csv")
        # Set when source_type=roster_xlsx and a roster has been parsed.
        self.std_canonical = None

        # LC state
        self.lc_si_df = None
        self.lc_so_df = None
        self.lc_use_signin_only = tk.BooleanVar(value=False)  # sign-in only mode
        # Roster mode: participants precomputed by the BSC-Manager web app
        # ("Export roster (.xlsx)" on /admin/ceu/:collaborativeId). When set,
        # Qualtrics CSV parsing + build_lc_attendance are skipped entirely and
        # the roster's hours are used verbatim.
        self.lc_roster_participants = None
        self.lc_session_rows = []  # list of (BooleanVar, date_str, hours_Entry)
        self._cancel_send = False
        self._lc_cancel   = False

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)
        _std_frame = ttk.Frame(self.notebook)
        self.notebook.add(_std_frame, text="  Standard Training  ")
        self.scroll = ScrollableFrame(_std_frame)
        self.scroll.pack(fill="both", expand=True)
        root = self.scroll.frame

        # ----- Event details -----
        frm1 = ttk.LabelFrame(root, text="Event Details")
        frm1.pack(fill="x", padx=12, pady=(12,6))

        ttk.Label(frm1, text="Training Event Title:").grid(row=0, column=0, sticky="w", padx=(8,6), pady=4)
        self.training = ttk.Entry(frm1, width=56); self.training.grid(row=0, column=1, sticky="w", padx=8, pady=4)

        ttk.Label(frm1, text="Date(s):").grid(row=1, column=0, sticky="nw", padx=(8,6), pady=4)
        date_frame = ttk.Frame(frm1); date_frame.grid(row=1, column=1, sticky="w", padx=8, pady=4)
        if DateEntry is not None:
            self.date_picker = DateEntry(date_frame, width=18, date_pattern="mm/dd/yyyy")
        else:
            self.date_picker = ttk.Entry(date_frame, width=22)
        self.date_picker.pack(side="left")
        ttk.Button(date_frame, text="Add Date", command=self._add_date).pack(side="left", padx=8)
        ttk.Button(date_frame, text="Remove Selected", command=self._remove_selected_date).pack(side="left", padx=8)
        self.date_list = tk.Listbox(frm1, height=3)
        self.date_list.grid(row=2, column=1, sticky="we", padx=8, pady=(0,6))

        self.location = self._labeled_entry(frm1, "Location:", 3)
        self.trainer  = self._labeled_entry(frm1, "Trainer:", 4)
        self.hours    = self._labeled_entry(frm1, "Hours:", 5); self.hours.insert(0, "3")

        # ----- Approved CEUs -----
        frm_st_ceu = ttk.LabelFrame(root, text="Approved CEUs  (select all that apply)")
        frm_st_ceu.pack(fill="x", padx=12, pady=6)

        self.st_ceu_sw    = tk.BooleanVar()
        self.st_ceu_psych = tk.BooleanVar()
        self.st_ceu_eila  = tk.BooleanVar()
        self.st_ceu_frsky = tk.BooleanVar()
        self.st_ceu_aswb  = tk.BooleanVar()
        self.st_ceu_lpcc  = tk.BooleanVar()

        ttk.Checkbutton(frm_st_ceu, text="KY Board of Social Work", variable=self.st_ceu_sw   ).grid(row=0, column=0, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_st_ceu, text="KY Board of Psychology",  variable=self.st_ceu_psych).grid(row=0, column=1, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_st_ceu, text="LPCC",                    variable=self.st_ceu_lpcc ).grid(row=0, column=2, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_st_ceu, text="ASWB",                    variable=self.st_ceu_aswb ).grid(row=1, column=0, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_st_ceu, text="EILA",  variable=self.st_ceu_eila,
            command=self._st_ceu_toggle).grid(row=1, column=1, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_st_ceu, text="FRSKY", variable=self.st_ceu_frsky,
            command=self._st_ceu_toggle).grid(row=1, column=2, sticky="w", padx=10, pady=3)

        self._st_ceu_cond_frm = ttk.Frame(frm_st_ceu)
        self._st_ceu_cond_frm.grid(row=2, column=0, columnspan=4, sticky="w", padx=6, pady=(0,6))

        self._st_ceu_eila_lbl  = ttk.Label(self._st_ceu_cond_frm, text="EILA #:")
        self.st_eila           = ttk.Entry(self._st_ceu_cond_frm, width=20)
        self._st_ceu_frsky_lbl = ttk.Label(self._st_ceu_cond_frm, text="FRSKY #:")
        self.st_frsky          = ttk.Entry(self._st_ceu_cond_frm, width=20)
        self.st_ceu_aswb.trace_add("write", lambda *_: self._st_ceu_toggle())
        self._st_ceu_aswb_lbl  = ttk.Label(self._st_ceu_cond_frm, text="ASWB Approval Period:")
        self.st_aswb_period    = ttk.Entry(self._st_ceu_cond_frm, width=30)

        self._st_ceu_toggle()   # initial state

        # ----- Sources -----
        frm2 = ttk.LabelFrame(root, text="Sources")
        frm2.pack(fill="x", padx=12, pady=6)

        # Source type — controls whether the Sign-In picker accepts a
        # Qualtrics CSV (existing flow) or a per-participant roster XLSX.
        ttk.Label(frm2, text="Source type:").grid(
            row=0, column=0, sticky="w", padx=(8,4), pady=(6,2))
        ttk.Radiobutton(
            frm2, text="Qualtrics CSV (sign-in/out)",
            variable=self.std_source_type, value="qualtrics_csv",
            command=self._std_on_source_change
        ).grid(row=0, column=1, sticky="w", padx=2)
        ttk.Radiobutton(
            frm2, text="Per-participant roster (XLSX)",
            variable=self.std_source_type, value="roster_xlsx",
            command=self._std_on_source_change
        ).grid(row=0, column=2, sticky="w", padx=2)
        self.std_source_hint = ttk.Label(frm2, text="", foreground="#666")
        self.std_source_hint.grid(row=1, column=0, columnspan=4, sticky="w",
                                  padx=(10,4), pady=(0, 2))

        # Method radio (Qualtrics flow only — ignored when roster source).
        self.method = tk.IntVar(value=2)  # 1 both, 2 sign-in only, 3 sign-out only
        ttk.Radiobutton(frm2, text="Use both sign-in & sign-out", variable=self.method, value=1).grid(row=2, column=0, sticky="w")
        ttk.Radiobutton(frm2, text="Use sign-in only", variable=self.method, value=2).grid(row=2, column=1, sticky="w")
        ttk.Radiobutton(frm2, text="Use sign-out only", variable=self.method, value=3).grid(row=2, column=2, sticky="w")
        # Sign-In picker accepts CSV OR XLSX so one row serves both source types.
        self.signin   = self._file_row(frm2, "Upload Sign-In / Roster:", 3,
                                       filetypes=[("CSV or XLSX", "*.csv *.xlsx"),
                                                  ("CSV", "*.csv"), ("XLSX", "*.xlsx")],
                                       on_pick=self._on_pick_signin)
        self.signout  = self._file_row(frm2, "Upload Sign Out:", 4, on_pick=self._on_pick_signout)
        self.template = self._file_row(frm2, "Upload Certificate Template:", 5, filetypes=[("Word", "*.docx")])
        self.outdir   = self._folder_row(frm2, "Select Folder to Save Certificates:", 6)
        # Initial hint state
        self._std_on_source_change()

        # ----- Participant Email -----
        frm3 = ttk.LabelFrame(root, text="Participant Email")
        frm3.pack(fill="both", expand=False, padx=12, pady=6)
        self.msg = tk.Text(frm3, height=7, wrap="word")
        self.msg.pack(fill="both", expand=True, padx=8, pady=8)
        self.msg.insert("1.0", DEFAULT_CERT_EMAIL)

        # ----- Single Certificate -----
        frm_single = ttk.LabelFrame(root, text="Single Certificate")
        frm_single.pack(fill="x", padx=12, pady=6)
        ttk.Label(frm_single, text="Participant Name:").grid(row=0, column=0, sticky="w", padx=10, pady=4)
        self.single_name = ttk.Entry(frm_single, width=36); self.single_name.grid(row=0, column=1, sticky="w", padx=6, pady=4)
        ttk.Label(frm_single, text="Participant Email:").grid(row=0, column=2, sticky="w", padx=10, pady=4)
        self.single_email = ttk.Entry(frm_single, width=36); self.single_email.grid(row=0, column=3, sticky="w", padx=6, pady=4)
        ttk.Button(frm_single, text="Send Single Certificate", command=self.send_single).grid(row=0, column=4, padx=10, pady=4)

        # ----- Actions -----
        frm4 = ttk.Frame(root); frm4.pack(fill="x", padx=12, pady=6)
        ttk.Button(frm4, text="Send Test Certificate", command=self.send_test).pack(side="left")
        ttk.Button(frm4, text="Send CEUs",   command=self.send_ceus).pack(side="left", padx=(8,0))
        ttk.Button(frm4, text="Cancel Send", command=self._cancel_std_send).pack(side="left", padx=(8,0))
        ttk.Button(frm4, text="Start Over",  command=self.start_over).pack(side="left", padx=8)

        # ----- Progress -----
        frm5 = ttk.LabelFrame(root, text="Progress")
        frm5.pack(fill="x", padx=12, pady=(6,12))
        self.progress = ttk.Progressbar(frm5, mode="determinate", maximum=100)
        self.progress.pack(fill="x", padx=10, pady=8)
        self.prog_label = ttk.Label(frm5, text="Idle.")
        self.prog_label.pack(anchor="w", padx=12, pady=(0,8))

        # ----- Invoice Request -----
        self._build_invoice_request(root)

        # ----- Update Training Spreadsheet -----
        self._build_training_spreadsheet(root)

        # ---- Learning Collaborative tab ----
        _lc_frame = ttk.Frame(self.notebook)
        self.notebook.add(_lc_frame, text="  Learning Collaborative  ")
        self.lc_scroll = ScrollableFrame(_lc_frame)
        self.lc_scroll.pack(fill="both", expand=True)
        self._build_lc_tab(self.lc_scroll.frame)

        # ---- Evaluation Report tab ----
        _eval_frame = ttk.Frame(self.notebook)
        self.notebook.add(_eval_frame, text="  Evaluation Report  ")
        self.eval_scroll = ScrollableFrame(_eval_frame)
        self.eval_scroll.pack(fill="both", expand=True)
        self._eval_df = None
        self._eval_session_rows = []
        self._build_eval_report_tab(self.eval_scroll.frame)

        # ---- Training Request tab ----
        _req_frame = ttk.Frame(self.notebook)
        self.notebook.add(_req_frame, text="  Training Request  ")
        self.req_scroll = ScrollableFrame(_req_frame)
        self.req_scroll.pack(fill="both", expand=True)
        # JSON store dir: training_requests/ next to the script
        self._req_store = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "training_requests")
        self._req_current_id = None   # ID of loaded request (None = new)
        self._build_request_tab(self.req_scroll.frame)


    # ---- tiny UI helpers ----
    def _labeled_entry(self, parent, label, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8,6), pady=4)
        e = ttk.Entry(parent, width=56)
        e.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        return e

    def _file_row(self, parent, label, row, filetypes=[("CSV", "*.csv")], on_pick=None):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8,6), pady=4)
        v = ttk.Entry(parent, width=56)
        v.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        def browse():
            p = filedialog.askopenfilename(title=label, filetypes=filetypes)
            if p:
                v.delete(0, tk.END); v.insert(0, p)
                if on_pick: on_pick(p)
        ttk.Button(parent, text="Browse…", command=browse).grid(row=row, column=2, padx=6)
        return v

    def _folder_row(self, parent, label, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8,6), pady=4)
        v = ttk.Entry(parent, width=56)
        v.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        def browse():
            p = filedialog.askdirectory(title=label)
            if p:
                v.delete(0, tk.END); v.insert(0, p)
        ttk.Button(parent, text="Browse…", command=browse).grid(row=row, column=2, padx=6)
        return v

    # ---- dates management ----
    def _add_date(self):
        d = self.date_picker.get().strip()
        if not d: return
        if d not in self.event_dates:
            self.event_dates.append(d)
            self._refresh_date_list()

    def _remove_selected_date(self):
        try:
            idx = self.date_list.curselection()
            if not idx: return
            sel = idx[0]
            val = self.date_list.get(sel)
            if val in self.event_dates:
                self.event_dates.remove(val)
                self._refresh_date_list()
        except Exception:
            pass

    def _refresh_date_list(self):
        self.date_list.delete(0, tk.END)
        for d in self.event_dates:
            self.date_list.insert(tk.END, d)

    # ---- event snapshot ----
    def _date_string(self):
        return format_dates_short(self.event_dates) if self.event_dates else ""

    def get_event(self):
        return {
            "TRAINING": self.training.get().strip(),
            "DATE": self._date_string(),  # multi-date formatted
            "LOCATION": self.location.get().strip(),
            "TRAINER": self.trainer.get().strip(),
            "HOURS": self.hours.get().strip(),
        }

    # ---- load sources ----
    # ---- load sources ----
    def _std_is_roster(self) -> bool:
        """True when the Standard tab is configured for a per-participant
        roster (XLSX). When True, the sign-out picker and method radios are
        unused and the upload flows through ingest_attendance()."""
        return self.std_source_type.get() == "roster_xlsx"

    def _std_on_source_change(self):
        """Update the on-screen hint and clear any cached parses when the
        Standard tab's Source type radio toggles."""
        if self._std_is_roster():
            self.std_source_hint.config(
                text="Roster mode: upload an XLSX with Name, Email, and per-session columns. "
                     "Sign-Out and method radios are ignored. Hours field is hours per session.")
        else:
            self.std_source_hint.config(text="")
        # Toss any cached parse — user is switching sources
        self.si_df = None
        self.so_df = None
        self.std_canonical = None
        self._update_participant_count_label()

    def ensure_sources_loaded(self):
        if self._std_is_roster():
            if not self.signin.get():
                raise ValueError("Please select a Roster XLSX (upload row).")
            if self.std_canonical is None or self.si_df is None:
                self._on_pick_signin(self.signin.get())
            if self.std_canonical is None or self.si_df is None:
                raise ValueError("Could not load the roster file.")
            return
        # Qualtrics path (unchanged)
        method = self.method.get()
        if method in (1,2):
            if not self.signin.get(): raise ValueError("Please select a Sign-In CSV.")
            if self.si_df is None: self.si_df = load_csv(self.signin.get())
        if method in (1,3):
            if not self.signout.get(): raise ValueError("Please select a Sign Out CSV.")
            if self.so_df is None: self.so_df = load_csv(self.signout.get())

    def _on_pick_signin(self, path):
        self.si_df = None
        self.std_canonical = None
        try:
            if self._std_is_roster():
                ca = ingest_attendance(path, profile=PerParticipantRosterProfile())
                self.std_canonical = ca
                # Build a compatibility si_df: one row per unique participant.
                # Downstream Qualtrics code paths expect NAME/EMAIL/ROLE columns.
                if ca.rows.empty:
                    self.si_df = pd.DataFrame(columns=["NAME","EMAIL","ROLE"])
                else:
                    df = (ca.rows.drop_duplicates(subset=["EMAIL"], keep="first")
                                  [["NAME","EMAIL"]].copy())
                    df = df[df["EMAIL"].apply(looks_like_email)]
                    df["ROLE"] = ""
                    self.si_df = df.reset_index(drop=True)
                if ca.warnings or ca.session_columns:
                    parts = [f"Loaded {len(self.si_df)} participants from "
                             f"{os.path.basename(path)}."]
                    if ca.session_columns:
                        parts.append("Session columns: "
                                     + ", ".join(s[0] for s in ca.session_columns))
                    if ca.warnings:
                        parts.append("Notes:\n  • " + "\n  • ".join(ca.warnings))
                    messagebox.showinfo("Roster Loaded", "\n\n".join(parts))
            else:
                self.si_df = load_csv(path)
        except Exception as e:
            messagebox.showerror("Sign-In / Roster", str(e))
            self.si_df = None
            self.std_canonical = None
        self._update_participant_count_label()

    def _on_pick_signout(self, path):
        self.so_df = None

    def _update_participant_count_label(self):
        count = len(self.si_df) if isinstance(self.si_df, pd.DataFrame) else 0
        self.participant_count.set(count)
        if hasattr(self, "tu_participants_label"):
            self.tu_participants_label.config(text=f"Participants (from Sign-In): {count}")

    # ---- buttons ----

    def start_over(self):
        for w in (self.training, self.location, self.trainer, self.hours, self.signin, self.signout, self.template, self.outdir,
                  self.single_name, self.single_email):
            try: w.delete(0, tk.END)
            except Exception: pass
        self.event_dates = []; self._refresh_date_list()
        if DateEntry is None:
            try: self.date_picker.delete(0, tk.END)
            except Exception: pass
        self.msg.delete("1.0", tk.END); self.msg.insert("1.0", DEFAULT_CERT_EMAIL)
        self.progress["value"] = 0; self.prog_label.config(text="Idle.")
        self._cancel_send = False
        self.si_df = self.so_df = None; self.participant_count.set(0)
        # Reset Standard-tab source type to the default Qualtrics flow
        self.std_source_type.set("qualtrics_csv")
        self.std_canonical = None
        self._std_on_source_change()
        # Reset standard CEU fields
        for v in (self.st_ceu_sw, self.st_ceu_psych, self.st_ceu_eila,
                  self.st_ceu_frsky, self.st_ceu_aswb, self.st_ceu_lpcc):
            v.set(False)
        for w in (self.st_eila, self.st_frsky, self.st_aswb_period):
            try: w.delete(0, tk.END)
            except Exception: pass
        self._st_ceu_toggle()
        # invoice reset
        self.inv_client.delete(0, tk.END); self.inv_contact.delete(0, tk.END); self.inv_email.delete(0, tk.END)
        self.inv_subject.delete(0, tk.END); self.inv_subject.insert(0, "Invoice needed")
        self._prefill_invoice_body()
        # spreadsheet reset (keep path)
        for key,e in self.mtl_entries.items():
            if isinstance(e, ttk.Combobox):
                e.set("")
            else:
                e.delete(0, tk.END)
        self.mtl_entries["CEU"].insert(0, "Yes")
        self.mtl_entries["Venue"].insert(0, "Zoom")

    def _cheer(self):
        try:
            if winsound:
                try:
                    winsound.MessageBeep(winsound.MB_ICONASTERISK)
                except Exception:
                    try: winsound.MessageBeep()
                    except Exception: self.bell()
            else:
                self.bell()
        except Exception:
            pass

    def send_test(self):
        try:
            event = self.get_event()
            if not self.template.get(): raise ValueError("Select a Certificate Template (DOCX).")
            outdir = self.outdir.get().strip() or os.path.join(os.getcwd(), "output")
            os.makedirs(outdir, exist_ok=True)
            ceu_vars = self._st_get_ceu_vars()
            fpath = generate_one(TEST_NAME, TEST_EMAIL, event, self.template.get(), outdir, ceu_vars=ceu_vars)
            subj = f"Certificate for {event['TRAINING']} , {event['DATE']} {ORG_NAME}"
            body_html = md_to_html(
                self.msg.get("1.0", tk.END)
                .replace("<<FIRSTNAME>>", extract_first(TEST_NAME))
                .replace("<<TRAINING>>", event["TRAINING"])
                .replace("<<LOCATION>>", event["LOCATION"])
            )
            send_outlook_email(TEST_EMAIL, subj, body_html, fpath, html=True)
            messagebox.showinfo("Test sent", f"Test certificate sent to {TEST_EMAIL}.")
        except Exception as e:
            messagebox.showerror("Test failed", str(e))

    def send_single(self):
        try:
            event = self.get_event()
            if not self.template.get(): raise ValueError("Select a Certificate Template (DOCX).")
            outdir = self.outdir.get().strip()
            if not outdir:
                raise ValueError("Please select the 'Select Folder to Save Certificates' path before sending a single certificate.")
            os.makedirs(outdir, exist_ok=True)
            name = smart_title(self.single_name.get().strip())
            email = self.single_email.get().strip()
            if not name: raise ValueError("Enter participant name.")
            if not looks_like_email(email): raise ValueError("Enter a valid participant email.")
            self.progress["value"] = 0; self.prog_label.config(text=f"Generating certificate for {name}…")
            self.update_idletasks()
            ceu_vars = self._st_get_ceu_vars()
            fpath = generate_one(name, email, event, self.template.get(), outdir, ceu_vars=ceu_vars)
            subj = f"Certificate for {event['TRAINING']} , {event['DATE']} {ORG_NAME}"
            body_html = md_to_html(
                self.msg.get("1.0", tk.END)
                .replace("<<FIRSTNAME>>", extract_first(name))
                .replace("<<TRAINING>>", event["TRAINING"])
                .replace("<<LOCATION>>", event["LOCATION"])
            )
            send_outlook_email(email, subj, body_html, fpath, html=True)
            self.prog_label.config(text=f"Sent to {name} <{email}>"); self._cheer()
            messagebox.showinfo("Done", f"Sent one certificate to {name}.")
        except Exception as e:
            messagebox.showerror("Single Certificate", str(e))

    def send_ceus(self):
        try:
            event = self.get_event()
            if not event["TRAINING"]: raise ValueError("Please enter Training Event Title.")
            if not self.template.get(): raise ValueError("Select a Certificate Template (DOCX).")
            outdir = self.outdir.get().strip() or os.path.join(os.getcwd(), "output")
            os.makedirs(outdir, exist_ok=True)
            self.ensure_sources_loaded()
            is_roster = self._std_is_roster() and self.std_canonical is not None
            # In roster mode the per-participant hours come from
            # sessions_attended × UI hours-per-session; method/signout are
            # ignored. We force sign-in-only matching so existing code paths
            # still produce the per-participant DataFrame we want.
            method = 2 if is_roster else self.method.get()

            def keys(df):
                if df is None or df.empty: return set()
                d = df.copy(); d["KEY"] = d["EMAIL"].str.lower() + d["NAME"].str.lower()
                return set(d["KEY"])

            so_df_for_match = None if is_roster else self.so_df
            final = match_participants(self.si_df, so_df_for_match, method)
            total = len(final)

            # Per-participant attendance counts for roster mode
            attendance_by_email = {}
            if is_roster:
                ca = self.std_canonical
                attendance_by_email = (
                    ca.rows.groupby(ca.rows["EMAIL"].str.lower()).size().to_dict()
                )
            self.progress["value"] = 0; self.progress["maximum"] = max(1, total)
            start_time = time.time()
            self.prog_label.config(text=f"Sending CEU 0 of {total}…"); self.update_idletasks()
            issued = []; fail_reasons = []
            msg_template = self.msg.get("1.0", tk.END)

            si_keys = keys(self.si_df); so_keys = keys(self.so_df if not is_roster else None); final_keys = keys(final)
            skipped_mismatch = []
            if method == 1:
                only_in  = sorted(si_keys - final_keys)
                only_out = sorted(so_keys - final_keys)
                for k in only_in:  skipped_mismatch.append((k, "Signed in but did not sign out"))
                for k in only_out: skipped_mismatch.append((k, "Signed out but did not sign in"))
            email_by_key = {}; name_by_key = {}
            for df in [self.si_df, None if is_roster else self.so_df]:
                if df is None: continue
                d = df.copy(); d["KEY"] = d["EMAIL"].str.lower() + d["NAME"].str.lower()
                for r in d.itertuples(index=False):
                    key = (getattr(r,"EMAIL").lower() + getattr(r,"NAME").lower())
                    email_by_key[key] = getattr(r,"EMAIL")
                    name_by_key[key] = getattr(r,"NAME")

            # Hours-per-session value entered in the UI — used in roster mode
            try:
                hours_per_session = float(str(event["HOURS"]).strip()) if event["HOURS"] else 0.0
            except (TypeError, ValueError):
                hours_per_session = 0.0

            self._cancel_send = False
            for i, row in enumerate(final.itertuples(index=False), start=1):
                if self._cancel_send:
                    self.prog_label.config(text="Send cancelled.")
                    break
                name = row.NAME; email = row.EMAIL
                if not looks_like_email(email):
                    fail_reasons.append((name, email, "Invalid email address")); continue
                # In roster mode, compute hours per participant before
                # generating the certificate.
                if is_roster:
                    sessions_attended = int(attendance_by_email.get(email.lower(), 0))
                    if sessions_attended <= 0:
                        fail_reasons.append((name, email, "No attendance rows found in roster"))
                        continue
                    this_hours = sessions_attended * hours_per_session
                    event_for_this = dict(event)
                    event_for_this["HOURS"] = f"{this_hours:g}"
                else:
                    event_for_this = event
                try:
                    ceu_vars = self._st_get_ceu_vars()
                    fpath = generate_one(name, email, event_for_this, self.template.get(), outdir, ceu_vars=ceu_vars)
                    subj = f"Certificate for {event['TRAINING']} , {event['DATE']} {ORG_NAME}"
                    body_html = md_to_html(
                        msg_template
                        .replace("<<FIRSTNAME>>", extract_first(name))
                        .replace("<<TRAINING>>", event["TRAINING"])
                        .replace("<<LOCATION>>", event["LOCATION"])
                    )
                    send_outlook_email(email, subj, body_html, fpath, html=True)
                    issued.append((name, email))
                except Exception as e:
                    fail_reasons.append((name, email, f"Email/Generate error: {e}"))

                elapsed = time.time() - start_time
                avg = elapsed / max(1, i); remaining = max(0, total - i)
                eta = int(avg * remaining); eta_str = f"{eta//60:02d}:{eta%60:02d}"
                self.progress["value"] = i
                self.prog_label.config(text=f"Sending CEU {i} of {total}…  ETA ~ {eta_str}")
                self.update_idletasks()

            for key, reason in skipped_mismatch:
                fail_reasons.append((name_by_key.get(key,"(unknown)"), email_by_key.get(key,""), reason))

            # Report email
            safe_training = safe_filename(event['TRAINING'])
            safe_date = safe_filename(event['DATE'])
            report_csv = os.path.join(outdir, f"Certificate Report - {safe_training} - {safe_date}.csv")
            self._send_completion_report(event, issued, fail_reasons, report_csv)

            messagebox.showinfo("All done", "All certificates processed and the report email was sent.")
            self.prog_label.config(text="Done.")
            self._update_participant_count_label()
            self._cheer()
        except Exception as e:
            messagebox.showerror("Send CEUs", str(e))

    def _send_completion_report(self, event, issued, fails, report_csv_path):
        try:
            os.makedirs(os.path.dirname(report_csv_path), exist_ok=True)
            issued_sorted = sorted(issued, key=lambda x: x[0].lower())
            with open(report_csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Issued Name","Issued Email"])
                for n,e in issued_sorted: w.writerow([n,e])
                w.writerow([]); w.writerow(["Not Issued Name","Email","Reason"])
                for n,e,r in fails: w.writerow([n,e,r])

            if not win32: raise RuntimeError("Outlook not available to send report.")
            body_lines = [f"Here is the certificate report for {event['TRAINING']} on {event['DATE']}", "", "Certificates Issued"]
            for i,(n,e) in enumerate(issued_sorted, start=1):
                body_lines.append(f"{i}. {n} <{e}>")
            body_lines.append("")
            if fails:
                body_lines.append("Certificates not issued")
                for n,e,r in fails:
                    email_txt = f" <{e}>" if e else ""
                    body_lines.append(f"- {n}{email_txt} — {r}")
            else:
                body_lines.append("Certificates not issued"); body_lines.append("- None")

            subj = f"Certificate report for {event['TRAINING']} , {event['DATE']}"
            body_html = md_to_html("\n".join(body_lines))
            send_outlook_email(TEST_EMAIL, subj, body_html, report_csv_path, html=True)
        except Exception as e:
            messagebox.showwarning("Report Email", f"Report email failed: {e}")

    # ----- Invoice Request -----

    def _build_invoice_request(self, root):
        frm = ttk.LabelFrame(root, text="Invoice Request")
        frm.pack(fill="both", expand=False, padx=12, pady=6)
        ttk.Label(frm, text="Client:").grid(row=0, column=0, sticky="w", padx=10, pady=4)
        self.inv_client = ttk.Entry(frm, width=60); self.inv_client.grid(row=0, column=1, sticky="w", padx=6, pady=4)
        ttk.Label(frm, text="Client Contact:").grid(row=1, column=0, sticky="w", padx=10, pady=4)
        self.inv_contact = ttk.Entry(frm, width=60); self.inv_contact.grid(row=1, column=1, sticky="w", padx=6, pady=4)
        ttk.Label(frm, text="Client Email:").grid(row=2, column=0, sticky="w", padx=10, pady=4)
        self.inv_email = ttk.Entry(frm, width=60); self.inv_email.grid(row=2, column=1, sticky="w", padx=6, pady=4)

        ttk.Label(frm, text="Subject:").grid(row=3, column=0, sticky="w", padx=10, pady=(10,2))
        self.inv_subject = ttk.Entry(frm, width=70); self.inv_subject.grid(row=3, column=1, sticky="w", padx=6, pady=(10,2))
        self.inv_subject.insert(0, "Invoice needed")

        ttk.Label(frm, text="Message:").grid(row=4, column=0, sticky="nw", padx=10, pady=(2,2))
        self.inv_msg = tk.Text(frm, height=6, wrap="word")
        self.inv_msg.grid(row=4, column=1, sticky="nsew", padx=6, pady=(2,8))
        frm.grid_rowconfigure(4, weight=1); frm.grid_columnconfigure(1, weight=1)

        btns = ttk.Frame(frm); btns.grid(row=5, column=1, sticky="e", padx=6, pady=(0,10))
        ttk.Button(btns, text="Prefill from Event", command=self._prefill_invoice_body).pack(side="left", padx=6)
        ttk.Button(btns, text="Send Invoice Request", command=self._send_invoice_request).pack(side="left", padx=6)

        self._prefill_invoice_body()

    def _prefill_invoice_body(self):
        ev = self.get_event()
        client  = self.inv_client.get().strip() or "(client)"
        contact = self.inv_contact.get().strip() or "(contact)"
        cemail  = self.inv_email.get().strip() or "(email)"
        default_body = (
            f"Hello Trish!\n\n"
            f"Letting you know that we have delivered {ev['TRAINING']} on {ev['DATE']} for ({client}). "
            f"If an invoice hasn’t already been sent, please send an invoice to ({contact}) at ({cemail}). "
            f"Please cc me and Dr. Sprang when you send. Thanks! Josh\n"
        )
        self.inv_msg.delete("1.0", tk.END); self.inv_msg.insert("1.0", default_body)

    def _send_invoice_request(self):
        try:
            if not win32: raise RuntimeError("Outlook not available.")
            subj = self.inv_subject.get().strip() or "Invoice needed"
            html = md_to_html(self.inv_msg.get("1.0", tk.END))
            send_outlook_email("trish.polly@uky.edu", subj, html, None, html=True, cc="sprang@uky.edu")
            messagebox.showinfo("Invoice", "Invoice request sent to Trish.")
        except Exception as e:
            messagebox.showerror("Invoice", str(e))

    # ----- Update Training Spreadsheet -----

    def _build_training_spreadsheet(self, root):
        frm = ttk.LabelFrame(root, text="Update Training Spreadsheet")
        frm.pack(fill="both", expand=False, padx=12, pady=6)
        row = 0

        # Entries (some are combobox)
        self.mtl_entries = {}

        def add_row(label, key, widget="entry", values=None):
            nonlocal row
            ttk.Label(frm, text=label).grid(row=row, column=0, sticky="w", padx=10, pady=4)
            if widget == "combo":
                cb = ttk.Combobox(frm, width=58)
                cb["values"] = values or []
                cb.grid(row=row, column=1, sticky="w", padx=10, pady=4)
                self.mtl_entries[key] = cb
            elif widget == "autocomplete":
                cb = AutoCompleteCombobox(frm, width=58)
                cb.set_completion_list(values or [])
                cb.grid(row=row, column=1, sticky="w", padx=10, pady=4)
                self.mtl_entries[key] = cb
            else:
                e = ttk.Entry(frm, width=60)
                e.grid(row=row, column=1, sticky="w", padx=10, pady=4)
                self.mtl_entries[key] = e
            row += 1

        add_row("Date", "Date")
        add_row("Program", "Program", widget="combo", values=PROGRAM_CHOICES)
        add_row("Training", "Training")
        add_row("Trainers", "Trainers")
        add_row("Training Length", "Training Length")
        add_row("Attendance", "Attendance")
        add_row("Psych Attendance", "Psych Attendance")
        add_row("Target Audience", "Target Audience")
        add_row("Evals/Outcome", "Evals/Outcome")
        add_row("CEU", "CEU")
        add_row("City", "City")
        add_row("County", "County", widget="autocomplete", values=KY_COUNTIES)
        add_row("Venue", "Venue")

        # Defaults
        self.mtl_entries["CEU"].insert(0, "Yes")
        self.mtl_entries["Venue"].insert(0, "Zoom")

        btns = ttk.Frame(frm); btns.grid(row=row, column=1, sticky="e", padx=10, pady=10)
        ttk.Button(btns, text="Prefill from Event",
                   command=self._prefill_master_from_event).pack(side="left", padx=6)
        ttk.Button(btns, text="Log to Master Training List",
                   command=self._add_to_master_list).pack(side="left", padx=6)
        row += 1

        # Path row — compact, below the main fields, for cases when the network
        # share isn't reachable (e.g. working from home with a local copy).
        path_frm = ttk.Frame(frm)
        path_frm.grid(row=row, column=0, columnspan=3, sticky="w", padx=8, pady=(2, 8))
        ttk.Label(path_frm, text="Master Training List path:",
                  foreground="grey").pack(side="left", padx=(2, 4))
        self.mtl_path = ttk.Entry(path_frm, width=52)
        self.mtl_path.pack(side="left")
        self.mtl_path.insert(0, MASTER_TRAINING_DEFAULT)
        ttk.Button(path_frm, text="Browse…",
                   command=self._browse_spreadsheet).pack(side="left", padx=4)

    def _browse_spreadsheet(self):
        p = filedialog.askopenfilename(
            title="Choose Master Training List",
            filetypes=[("Excel Workbook","*.xlsx"),("All files","*.*")]
        )
        if p:
            self.mtl_path.delete(0, tk.END)
            self.mtl_path.insert(0, p)

    def _prefill_master_from_event(self):
        ev = self.get_event()
        if self.si_df is None and self.signin.get():
            try: self.si_df = load_csv(self.signin.get())
            except Exception: pass
        count = len(self.si_df) if isinstance(self.si_df, pd.DataFrame) else 0
        psych = count_psychologists(self.si_df)

        self.mtl_entries["Date"].delete(0, tk.END); self.mtl_entries["Date"].insert(0, ev["DATE"])
        self.mtl_entries["Training"].delete(0, tk.END); self.mtl_entries["Training"].insert(0, ev["TRAINING"])
        self.mtl_entries["Trainers"].delete(0, tk.END); self.mtl_entries["Trainers"].insert(0, ev["TRAINER"])
        self.mtl_entries["Training Length"].delete(0, tk.END); self.mtl_entries["Training Length"].insert(0, str(ev["HOURS"]))
        self.mtl_entries["Attendance"].delete(0, tk.END); self.mtl_entries["Attendance"].insert(0, str(count))
        self.mtl_entries["Psych Attendance"].delete(0, tk.END); self.mtl_entries["Psych Attendance"].insert(0, str(psych))

    def _add_to_master_list(self):
        """Append one row to the Training Information tab of the Master Training List."""
        try:
            path = self.mtl_path.get().strip()
            if not path:
                raise ValueError("Master Training List path is blank.")
            directory = os.path.dirname(path)
            if directory and not os.path.isdir(directory):
                raise ValueError(
                    "Path not reachable:\n" + path +
                    "\n\nIf this is a network share, make sure you're connected to it.")

            # Auto-fill Attendance / Psych from loaded sign-in if fields are blank
            if self.si_df is None and self.signin.get():
                try: self.si_df = load_csv(self.signin.get())
                except Exception: pass
            count = len(self.si_df) if isinstance(self.si_df, pd.DataFrame) else 0
            psych = count_psychologists(self.si_df)

            headers = [
                "Date","Program","Training","Trainers","Training Length","Attendance",
                "Psych Attendance","Target Audience","Evals/Outcome","CEU","City","County","Venue"
            ]
            row_vals = []
            for h in headers:
                w = self.mtl_entries[h]
                val = w.get().strip()
                if h == "Attendance"      and not val: val = str(count) if count else ""
                if h == "Psych Attendance" and not val: val = str(psych)  if count else ""
                row_vals.append(val)

            if not path.lower().endswith(".xlsx"):
                raise RuntimeError("The Master Training List must be an .xlsx file.")

            append_row_openpyxl(path, "Training Information", row_vals)
            messagebox.showinfo("Master Training List",
                "Row added to 'Training Information' tab successfully.")
        except Exception as e:
            messagebox.showerror("Master Training List", str(e))

    # ================================================================
    # Learning Collaborative tab
    # ================================================================

    def _build_lc_tab(self, root):
        # ----- Event Details -----
        frm1 = ttk.LabelFrame(root, text="Event Details")
        frm1.pack(fill="x", padx=12, pady=(12, 6))
        ttk.Label(frm1, text="Training Name:").grid(row=0, column=0, sticky="w", padx=(8,6), pady=4)
        self.lc_training = ttk.Entry(frm1, width=56)
        self.lc_training.grid(row=0, column=1, sticky="w", padx=8, pady=4)
        ttk.Label(frm1, text="Trainer:").grid(row=1, column=0, sticky="w", padx=(8,6), pady=4)
        self.lc_trainer = ttk.Entry(frm1, width=56)
        self.lc_trainer.grid(row=1, column=1, sticky="w", padx=8, pady=4)
        ttk.Label(frm1, text="Location:").grid(row=2, column=0, sticky="w", padx=(8,6), pady=4)
        self.lc_location = ttk.Entry(frm1, width=56)
        self.lc_location.grid(row=2, column=1, sticky="w", padx=8, pady=4)

        # ----- Approved CEUs -----
        frm_ceu = ttk.LabelFrame(root, text="Approved CEUs  (select all that apply)")
        frm_ceu.pack(fill="x", padx=12, pady=6)

        self.lc_ceu_sw    = tk.BooleanVar()
        self.lc_ceu_psych = tk.BooleanVar()
        self.lc_ceu_eila  = tk.BooleanVar()
        self.lc_ceu_frsky = tk.BooleanVar()
        self.lc_ceu_aswb  = tk.BooleanVar()
        self.lc_ceu_lpcc  = tk.BooleanVar()

        ttk.Checkbutton(frm_ceu, text="KY Board of Social Work",   variable=self.lc_ceu_sw   ).grid(row=0, column=0, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_ceu, text="KY Board of Psychology",    variable=self.lc_ceu_psych).grid(row=0, column=1, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_ceu, text="LPCC",                      variable=self.lc_ceu_lpcc ).grid(row=0, column=2, sticky="w", padx=10, pady=3)
        ttk.Checkbutton(frm_ceu, text="ASWB",                      variable=self.lc_ceu_aswb ).grid(row=1, column=0, sticky="w", padx=10, pady=3)

        # EILA checkbox + number
        ttk.Checkbutton(frm_ceu, text="EILA",  variable=self.lc_ceu_eila,
                        command=self._lc_ceu_toggle).grid(row=1, column=1, sticky="w", padx=10, pady=3)
        # FRSKY checkbox + number
        ttk.Checkbutton(frm_ceu, text="FRSKY", variable=self.lc_ceu_frsky,
                        command=self._lc_ceu_toggle).grid(row=1, column=2, sticky="w", padx=10, pady=3)

        # Conditional fields row
        self._ceu_cond_frm = ttk.Frame(frm_ceu)
        self._ceu_cond_frm.grid(row=2, column=0, columnspan=4, sticky="w", padx=6, pady=(0,6))

        # EILA # field
        self._ceu_eila_lbl = ttk.Label(self._ceu_cond_frm, text="EILA #:")
        self.lc_eila = ttk.Entry(self._ceu_cond_frm, width=20)

        # FRSKY # field
        self._ceu_frsky_lbl = ttk.Label(self._ceu_cond_frm, text="FRSKY #:")
        self.lc_frsky = ttk.Entry(self._ceu_cond_frm, width=20)

        # ASWB period field
        self.lc_ceu_aswb.trace_add("write", lambda *_: self._lc_ceu_toggle())
        self._ceu_aswb_lbl = ttk.Label(self._ceu_cond_frm, text="ASWB Approval Period:")
        self.lc_aswb_period = ttk.Entry(self._ceu_cond_frm, width=30)

        self._lc_ceu_toggle()   # initial state

        # ----- Sources -----
        frm2 = ttk.LabelFrame(root, text="Sources")
        frm2.pack(fill="x", padx=12, pady=6)
        self.lc_signin_path  = self._lc_file_row(frm2, "Sign-In CSV:",            0, on_pick=self._lc_on_pick_signin)
        self.lc_signout_path = self._lc_file_row(frm2, "Sign-Out CSV:",           1, on_pick=self._lc_on_pick_signout)
        # Roster mode: precomputed Name/Email/Hours roster exported by the
        # BSC-Manager web app. When picked, sign-in/out CSVs and Detect
        # Sessions are not required — hours come from the roster verbatim.
        self.lc_roster_path  = self._lc_file_row(frm2, "Precomputed Roster (App):", 2,
                                                 filetypes=[("Excel/CSV", "*.xlsx *.csv")],
                                                 on_pick=self._lc_on_pick_roster)
        ttk.Label(frm2, text="Date Range (roster mode):").grid(row=3, column=0, sticky="w", padx=(8,6), pady=4)
        self.lc_roster_daterange = ttk.Entry(frm2, width=30)
        self.lc_roster_daterange.grid(row=3, column=1, sticky="w", padx=8, pady=4)
        self.lc_template     = self._lc_file_row(frm2, "Certificate Template:",   4, filetypes=[("Word", "*.docx")])
        self.lc_outdir       = self._lc_folder_row(frm2, "Output Folder:",        5)
        ttk.Checkbutton(
            frm2,
            text="Use sign-in only (treat sign-in as attendance; sign-out CSV not required)",
            variable=self.lc_use_signin_only,
            command=self._lc_toggle_signin_only,
        ).grid(row=6, column=0, columnspan=4, sticky="w", padx=8, pady=(2, 2))
        ttk.Button(frm2, text="Detect Sessions", command=self._lc_detect_sessions).grid(
            row=7, column=1, sticky="w", padx=8, pady=(4, 8))

        # ----- Sessions -----
        frm3 = ttk.LabelFrame(root, text="Sessions  (uncheck any test or invalid dates)")
        frm3.pack(fill="x", padx=12, pady=6)
        ttk.Label(frm3, text="Include", width=8).grid(row=0, column=0, padx=(10,2), pady=4)
        ttk.Label(frm3, text="Date").grid(             row=0, column=1, padx=(2,4),  pady=4, sticky="w")
        ttk.Label(frm3, text="Hours").grid(            row=0, column=2, padx=(4,8),  pady=4, sticky="w")
        self.lc_sessions_frame = frm3
        self._lc_no_sessions_msg()


        # ----- Participant Email -----
        frm5 = ttk.LabelFrame(
            root, text="Participant Email  "
                       "(<<ATTENDANCE_TABLE>> is replaced with the per-session grid)")
        frm5.pack(fill="both", expand=False, padx=12, pady=6)
        self.lc_msg = tk.Text(frm5, height=9, wrap="word")
        self.lc_msg.pack(fill="both", expand=True, padx=8, pady=8)
        self.lc_msg.insert("1.0", DEFAULT_LC_EMAIL)

        # ----- Actions -----
        frm6 = ttk.Frame(root)
        frm6.pack(fill="x", padx=12, pady=6)
        ttk.Button(frm6, text="Send Test Certificate",  command=self._lc_send_test).pack(side="left")
        ttk.Button(frm6, text="Send All Certificates",  command=self._lc_review_before_send).pack(side="left", padx=(8,0))
        ttk.Button(frm6, text="Cancel Send",            command=self._lc_cancel_send).pack(side="left", padx=(8,0))
        ttk.Button(frm6, text="Start Over",             command=self._lc_start_over).pack(side="left", padx=8)

        # ----- Log Session to Master Training List -----
        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=12, pady=(10, 2))
        frm_log = ttk.LabelFrame(root,
            text="Log Session to Master Training List")
        frm_log.pack(fill="x", padx=12, pady=6)
        frm_log.columnconfigure(1, weight=1)

        # Session picker
        ttk.Label(frm_log, text="Session to Log:").grid(
            row=0, column=0, sticky="w", padx=(8,4), pady=4)
        self.lc_log_session_var = tk.StringVar()
        self.lc_log_session_cb  = ttk.Combobox(
            frm_log, textvariable=self.lc_log_session_var,
            state="readonly", width=36)
        self.lc_log_session_cb.grid(row=0, column=1, sticky="w", padx=8, pady=4)
        ttk.Button(frm_log, text="Prefill Fields",
                   command=self._lc_log_prefill).grid(
            row=0, column=2, padx=6, pady=4)

        # Data fields — matching Training Information columns
        self.lc_log_entries = {}

        def llog(label, key, r, widget="entry", values=None, width=40):
            ttk.Label(frm_log, text=label).grid(
                row=r, column=0, sticky="w", padx=(8,4), pady=3)
            if widget == "combo":
                w = ttk.Combobox(frm_log, values=values or [], width=width)
            elif widget == "autocomplete":
                w = AutoCompleteCombobox(frm_log, width=width)
                w.set_completion_list(values or [])
            else:
                w = ttk.Entry(frm_log, width=width+2)
            w.grid(row=r, column=1, sticky="w", padx=8, pady=3)
            self.lc_log_entries[key] = w

        llog("Training Name:",     "Training",        1, width=54)
        llog("Program:",           "Program",         2, widget="combo",
             values=PROGRAM_CHOICES, width=30)
        llog("Date:",              "Date",            3, width=20)
        llog("Hours:",             "Training Length", 4, width=10)
        llog("Attendance:",        "Attendance",      5, width=10)
        llog("Psych Attendance:",  "Psych Attendance",6, width=10)
        llog("Target Audience:",   "Target Audience", 7, width=54)
        llog("Evals/Outcome:",     "Evals/Outcome",   8, width=20)
        llog("CEU:",               "CEU",             9, width=30)
        llog("City:",              "City",           10, width=30)
        llog("County:",            "County",         11, widget="autocomplete",
             values=KY_COUNTIES, width=28)
        llog("Venue:",             "Venue",          12, width=30)

        # Defaults
        self.lc_log_entries["Evals/Outcome"].insert(0, "Yes")
        self.lc_log_entries["Venue"].insert(0, "Zoom")

        frm_log_btns = ttk.Frame(frm_log)
        frm_log_btns.grid(row=13, column=1, sticky="w", padx=8, pady=8)
        ttk.Button(frm_log_btns, text="Log This Session",
                   command=self._lc_log_session).pack(side="left", padx=4)
        ttk.Button(frm_log_btns, text="Log All Sessions",
                   command=self._lc_log_all_sessions).pack(side="left", padx=4)
        self.lc_log_status = ttk.Label(frm_log_btns, text="", foreground="green")
        self.lc_log_status.pack(side="left", padx=10)

        # ----- Progress -----
        frm7 = ttk.LabelFrame(root, text="Progress")
        frm7.pack(fill="x", padx=12, pady=(6, 12))
        self.lc_progress = ttk.Progressbar(frm7, mode="determinate", maximum=100)
        self.lc_progress.pack(fill="x", padx=10, pady=8)
        self.lc_prog_label = ttk.Label(frm7, text="Idle.")
        self.lc_prog_label.pack(anchor="w", padx=12, pady=(0, 8))

    # ---- LC tiny UI helpers ----

    def _st_ceu_toggle(self, *_):
        """Show/hide conditional CEU fields for Standard Training (EILA #, FRSKY #, ASWB period)."""
        col = 0
        for lbl, widget, var in [
            (self._st_ceu_eila_lbl,  self.st_eila,        self.st_ceu_eila),
            (self._st_ceu_frsky_lbl, self.st_frsky,        self.st_ceu_frsky),
            (self._st_ceu_aswb_lbl,  self.st_aswb_period,  self.st_ceu_aswb),
        ]:
            if var.get():
                lbl.grid(row=0, column=col, sticky="w", padx=(10,2), pady=2)
                widget.grid(row=0, column=col+1, sticky="w", padx=(0,16), pady=2)
                col += 2
            else:
                lbl.grid_remove()
                widget.grid_remove()

    def _st_get_ceu_vars(self) -> dict:
        """Return CEU placeholder dict for Standard Training certificate/email generation."""
        hours    = self.hours.get().strip() or "?"
        eila_no  = self.st_eila.get().strip()
        frsky_no = self.st_frsky.get().strip()
        period   = self.st_aswb_period.get().strip()

        sw_text    = CEU_SWBOARD_TEXT    if self.st_ceu_sw.get()    else ""
        psych_text = CEU_PSYCHBOARD_TEXT if self.st_ceu_psych.get() else ""
        lpcc_text  = CEU_LPCC_TEXT       if self.st_ceu_lpcc.get()  else ""
        eila_text  = CEU_EILA_TEMPLATE.format(hours=hours, number=eila_no)   if self.st_ceu_eila.get()  else ""
        frsky_text = CEU_FRSKY_TEMPLATE.format(hours=hours, number=frsky_no) if self.st_ceu_frsky.get() else ""
        aswb_text  = CEU_ASWB_TEMPLATE.format(period=period, hours=hours)    if self.st_ceu_aswb.get()  else ""

        all_approvals = "\n".join(t for t in [sw_text, psych_text, eila_text, frsky_text, aswb_text, lpcc_text] if t)

        return {
            "<<SWBOARD>>":       sw_text,
            "<<PSYCHBOARD>>":    psych_text,
            "<<LPCC>>":          lpcc_text,
            "<<EILA_APPROVAL>>": eila_text,
            "<<FRSKY>>":         frsky_text,
            "<<ASWB>>":          aswb_text,
            "<<CEU_APPROVALS>>": all_approvals,
            "<<TOTALHOURS>>":    hours,
        }

    def _cancel_std_send(self):
        self._cancel_send = True
        self.prog_label.config(text="Cancelling…")

    def _lc_cancel_send(self):
        self._lc_cancel = True
        self.lc_prog_label.config(text="Cancelling…")

    def _lc_ceu_toggle(self, *_):
        """Show/hide conditional CEU fields (EILA #, FRSKY # and ASWB period)."""
        col = 0
        for lbl, widget, var in [
            (self._ceu_eila_lbl,  self.lc_eila,        self.lc_ceu_eila),
            (self._ceu_frsky_lbl, self.lc_frsky,        self.lc_ceu_frsky),
            (self._ceu_aswb_lbl,  self.lc_aswb_period,  self.lc_ceu_aswb),
        ]:
            if var.get():
                lbl.grid(row=0, column=col, sticky="w", padx=(10,2), pady=2)
                widget.grid(row=0, column=col+1, sticky="w", padx=(0,16), pady=2)
                col += 2
            else:
                lbl.grid_remove()
                widget.grid_remove()

    def _lc_file_row(self, parent, label, row, filetypes=[("CSV", "*.csv")], on_pick=None):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8,6), pady=4)
        v = ttk.Entry(parent, width=56)
        v.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        def browse():
            p = filedialog.askopenfilename(title=label, filetypes=filetypes)
            if p:
                v.delete(0, tk.END); v.insert(0, p)
                if on_pick: on_pick(p)
        ttk.Button(parent, text="Browse…", command=browse).grid(row=row, column=2, padx=6)
        return v

    def _lc_folder_row(self, parent, label, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8,6), pady=4)
        v = ttk.Entry(parent, width=56)
        v.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        def browse():
            p = filedialog.askdirectory(title=label)
            if p:
                v.delete(0, tk.END); v.insert(0, p)
        ttk.Button(parent, text="Browse…", command=browse).grid(row=row, column=2, padx=6)
        return v

    def _lc_no_sessions_msg(self):
        ttk.Label(
            self.lc_sessions_frame,
            text="No sessions detected yet.  Upload both CSVs then click 'Detect Sessions'.",
            foreground="gray"
        ).grid(row=1, column=0, columnspan=4, padx=10, pady=6, sticky="w")

    def _lc_clear_session_rows(self):
        frm = self.lc_sessions_frame
        for widget in list(frm.winfo_children()):
            try:
                info = widget.grid_info()
                if info and int(info.get("row", 0)) >= 1:
                    widget.destroy()
            except Exception:
                pass
        self.lc_session_rows = []

    # ---- LC source picks ----

    def _lc_on_pick_signin(self, path):
        self.lc_si_df = None
        try:
            self.lc_si_df = parse_lc_signin_sessions(path)
        except Exception as e:
            messagebox.showerror("Sign-In CSV", str(e))

    def _lc_on_pick_signout(self, path):
        self.lc_so_df = None
        try:
            self.lc_so_df = parse_lc_signout_sessions(path)
        except Exception as e:
            messagebox.showerror("Sign-Out CSV", str(e))

    def _lc_on_pick_roster(self, path):
        """Load a precomputed roster exported by the BSC-Manager web app.

        Required columns: Name, Email, Hours Attended, Hours Total. Builds the
        participants list directly (no Qualtrics parsing, no
        build_lc_attendance); hours are used verbatim. session_data stays
        empty — the email's attendance table degrades to a one-line summary.
        """
        try:
            if path.lower().endswith(".xlsx"):
                df = pd.read_excel(path, dtype=object)
            else:
                df = pd.read_csv(path, dtype=object)
            df.columns = [str(c).strip() for c in df.columns]
            required = ["Name", "Email", "Hours Attended", "Hours Total"]
            missing = [c for c in required if c not in df.columns]
            if missing:
                raise ValueError(
                    "Roster is missing required column(s): " + ", ".join(missing)
                    + ". Expected exactly: Name, Email, Hours Attended, Hours Total.")
            participants = []
            for _, row in df.iterrows():
                name  = str(row["Name"]).strip()
                email = str(row["Email"]).strip()
                if not name or not looks_like_email(email):
                    continue
                participants.append({
                    "name":           name,
                    "email":          email,
                    "hours_attended": float(row["Hours Attended"] or 0),
                    "hours_total":    float(row["Hours Total"] or 0),
                    "session_data":   [],
                    "fuzzy_matched":  False,
                    "signin_only":    True,
                })
            if not participants:
                raise ValueError("No valid rows found in the roster.")
            self.lc_roster_participants = participants
            messagebox.showinfo(
                "Roster loaded",
                f"{len(participants)} participant(s) loaded from the app roster.\n\n"
                "Sign-in/out CSVs and Detect Sessions are not needed in roster "
                "mode. Hours will be used verbatim from the roster.")
        except Exception as e:
            self.lc_roster_participants = None
            messagebox.showerror("Roster import failed", str(e))

    def _lc_get_participants(self, sessions, excluded):
        """Single source for the participants list across review/test/send.

        Roster mode (precomputed app roster loaded): returns roster
        participants, applying only the staff-exclusion name filter. CSV mode:
        existing Qualtrics path via build_lc_attendance.
        """
        if self.lc_roster_participants is not None:
            excluded_lower = {e.strip().lower() for e in excluded if e.strip()}
            return [p for p in self.lc_roster_participants
                    if p["name"].strip().lower() not in excluded_lower]
        # ---- CSV mode (existing behavior) ----
        if self.lc_si_df is None:
            if not self.lc_signin_path.get():
                raise ValueError("Select a Sign-In CSV, or load a Precomputed Roster.")
            self.lc_si_df = parse_lc_signin_sessions(self.lc_signin_path.get())
        signin_only = self.lc_use_signin_only.get()
        if not signin_only and self.lc_so_df is None:
            if not self.lc_signout_path.get():
                raise ValueError("Select a Sign-Out CSV, or check 'Use sign-in only'.")
            self.lc_so_df = parse_lc_signout_sessions(self.lc_signout_path.get())
        so_for_build = None if signin_only else self.lc_so_df
        return build_lc_attendance(self.lc_si_df, so_for_build, sessions, excluded)

    def _lc_toggle_signin_only(self):
        """Called when the 'Use sign-in only' checkbox is toggled.
        Clears the cached sign-out DataFrame so downstream logic re-evaluates."""
        if self.lc_use_signin_only.get():
            # Entering sign-in-only mode — forget any previously loaded sign-out data
            self.lc_so_df = None

    # ---- LC session detection ----

    def _lc_detect_sessions(self):
        try:
            if not self.lc_signin_path.get():
                raise ValueError("Please upload the Sign-In CSV first.")
            if self.lc_si_df is None:
                self.lc_si_df = parse_lc_signin_sessions(self.lc_signin_path.get())
            signin_only = self.lc_use_signin_only.get()
            if not signin_only:
                if not self.lc_signout_path.get():
                    raise ValueError(
                        "Please upload the Sign-Out CSV first, "
                        "or check 'Use sign-in only' above.")
                if self.lc_so_df is None:
                    self.lc_so_df = parse_lc_signout_sessions(self.lc_signout_path.get())

            so_for_detect = None if signin_only else self.lc_so_df
            session_info = get_lc_session_dates(self.lc_si_df, so_for_detect)
            self._lc_clear_session_rows()

            if not session_info:
                self._lc_no_sessions_msg()
                return

            frm = self.lc_sessions_frame
            for i, (date_str, likely_real, count) in enumerate(session_info):
                var = tk.BooleanVar(value=likely_real)
                ttk.Checkbutton(frm, variable=var).grid(row=i+1, column=0, padx=(14,2), pady=2)
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    fmt = "%#m/%#d/%Y" if os.name == "nt" else "%-m/%-d/%Y"
                    date_display = dt.strftime(fmt)
                except Exception:
                    date_display = date_str
                lbl_txt = f"{date_display}   ({count} sign-in{'s' if count != 1 else ''})"
                ttk.Label(frm, text=lbl_txt, width=32).grid(row=i+1, column=1, padx=(2,4), pady=2, sticky="w")
                hours_e = ttk.Entry(frm, width=7)
                hours_e.insert(0, "3")
                hours_e.grid(row=i+1, column=2, padx=(4,8), pady=2)
                if not likely_real:
                    ttk.Label(frm, text="\u2190 likely test entry", foreground="#888").grid(
                        row=i+1, column=3, padx=4, pady=2, sticky="w")
                self.lc_session_rows.append((var, date_str, hours_e))

            real_count = sum(1 for _, r, _ in session_info if r)
            messagebox.showinfo(
                "Sessions Detected",
                f"Found {len(session_info)} date(s).  "
                f"{real_count} appear to be real sessions.\n\n"
                "Please review, set hours for each session, and uncheck any test entries.")
            # Also refresh the log-session dropdown
            self._lc_log_refresh_sessions()
        except Exception as e:
            messagebox.showerror("Detect Sessions", str(e))

    # ---- LC exclusion helpers ----

    def _lc_add_exclusion(self):
        name = self.lc_excl_entry.get().strip()
        if name and name not in list(self.lc_excl_listbox.get(0, tk.END)):
            self.lc_excl_listbox.insert(tk.END, name)
        self.lc_excl_entry.delete(0, tk.END)

    def _lc_remove_exclusion(self):
        sel = self.lc_excl_listbox.curselection()
        if sel: self.lc_excl_listbox.delete(sel[0])

    # ---- LC data getters ----

    def _lc_get_sessions(self) -> list:
        """Return list of (date_str, hours_float) for all checked sessions."""
        result = []
        for var, date_str, hours_e in self.lc_session_rows:
            if var.get():
                try:    h = float(hours_e.get().strip())
                except: h = 0.0
                result.append((date_str, h))
        return result

    def _lc_get_excluded(self) -> list:
        # Staff/trainers who should never receive certificates
        return ["Tracy Clemans", "Josh Fisherkeller", "Joshua Fisherkeller",
                "Leah Riggs", "Test"]

    def _lc_get_ceu_vars(self, sessions=None) -> dict:
        """Return a dict of CEU placeholder → text for cert/email generation."""
        fmt_n   = lambda n: f"{float(n):g}"
        if sessions:
            hours = fmt_n(sum(h for _, h in sessions))
        elif self.lc_roster_participants:
            # Roster mode: total hours come verbatim from the app's export.
            hours = fmt_n(self.lc_roster_participants[0]["hours_total"])
        else:
            hours = self.lc_eila.get().strip() or "?"
        eila_no = self.lc_eila.get().strip()
        frsky_no= self.lc_frsky.get().strip() if hasattr(self, "lc_frsky") else ""
        period  = self.lc_aswb_period.get().strip() if hasattr(self, "lc_aswb_period") else ""

        sw_text    = CEU_SWBOARD_TEXT    if self.lc_ceu_sw.get()    else ""
        psych_text = CEU_PSYCHBOARD_TEXT if self.lc_ceu_psych.get() else ""
        lpcc_text  = CEU_LPCC_TEXT       if self.lc_ceu_lpcc.get()  else ""
        eila_text  = CEU_EILA_TEMPLATE.format(hours=hours, number=eila_no) if self.lc_ceu_eila.get() else ""
        frsky_text = CEU_FRSKY_TEMPLATE.format(hours=hours, number=frsky_no) if self.lc_ceu_frsky.get() else ""
        aswb_text  = CEU_ASWB_TEMPLATE.format(period=period, hours=hours) if self.lc_ceu_aswb.get() else ""

        all_approvals = "\n".join(t for t in [sw_text, psych_text, eila_text, frsky_text, aswb_text, lpcc_text] if t)

        return {
            "<<SWBOARD>>":      sw_text,
            "<<PSYCHBOARD>>":   psych_text,
            "<<LPCC>>":         lpcc_text,
            "<<EILA_APPROVAL>>": eila_text,
            "<<FRSKY>>":        frsky_text,
            "<<ASWB>>":         aswb_text,
            "<<CEU_APPROVALS>>": all_approvals,
            "<<TOTALHOURS>>":   hours,
        }

    # ---- LC email body builder ----

    def _lc_build_email_html(self, participant: dict, event_name: str,
                              sessions: list) -> str:
        """Build personalized HTML email body for one LC participant."""
        msg_template = self.lc_msg.get("1.0", tk.END)
        first_name = extract_first(participant["name"])
        h_att = participant["hours_attended"]
        h_tot = participant["hours_total"]
        fmt_n = lambda n: f"{float(n):g}"

        # Format session dates for the attendance table
        dates_fmt = []
        for date_str, _ in sessions:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
                fmt = "%#m/%#d/%Y" if os.name == "nt" else "%-m/%-d/%Y"
                dates_fmt.append(dt.strftime(fmt))
            except Exception:
                dates_fmt.append(date_str)

        attendance_html = build_attendance_html(participant, dates_fmt)

        # Substitute text placeholders first
        ceu_vars = self._lc_get_ceu_vars(sessions)
        text = msg_template
        text = text.replace("<<FIRSTNAME>>",     first_name)
        text = text.replace("<<TRAINING>>",       event_name)
        text = text.replace("<<HOURS_ATTENDED>>", fmt_n(h_att))
        text = text.replace("<<HOURS_TOTAL>>",    fmt_n(h_tot))
        for ph, val in ceu_vars.items():
            text = text.replace(ph, val)

        # Convert plain text to HTML (escapes < > & and wraps newlines in <br>)
        html_body = md_to_html(text)

        # Replace the now-escaped attendance table placeholder with actual HTML
        html_body = html_body.replace("&lt;&lt;ATTENDANCE_TABLE&gt;&gt;", attendance_html)
        return html_body

    # ---- LC actions ----

    def _lc_review_before_send(self):
        """Show a review dialog listing who will/won't get certificates before sending."""
        try:
            event_name = self.lc_training.get().strip()
            if not event_name:
                raise ValueError("Enter Training Name.")
            sessions = self._lc_get_sessions()
            if not sessions and self.lc_roster_participants is None:
                raise ValueError("No sessions selected. Click 'Detect Sessions' first, or load a Precomputed Roster.")
            if not self.lc_template.get():
                raise ValueError("Select a Certificate Template (DOCX).")

            excluded     = self._lc_get_excluded()
            participants = self._lc_get_participants(sessions, excluded)

            will_send    = [p for p in participants if p["hours_attended"] > 0]
            wont_send    = [p for p in participants if p["hours_attended"] == 0]
            fuzzy_list   = [p for p in participants if p.get("fuzzy_matched")]
            fmt_n        = lambda n: f"{float(n):g}"

            # Build review window
            win = tk.Toplevel(self)
            win.title("Pre-Send Review")
            win.geometry("820x620")
            win.grab_set()

            ttk.Label(win, text=f"Review for: {event_name}",
                      font=("", 12, "bold")).pack(anchor="w", padx=14, pady=(12,4))
            ttk.Label(win,
                text=f"Will send: {len(will_send)}   Will NOT send: {len(wont_send)}"
                     + (f"   Fuzzy-matched names: {len(fuzzy_list)}" if fuzzy_list else ""),
                foreground="#333").pack(anchor="w", padx=14, pady=(0,8))

            pane = ttk.PanedWindow(win, orient="vertical")
            pane.pack(fill="both", expand=True, padx=10, pady=4)

            # ---- Will send ----
            frm_yes = ttk.LabelFrame(pane, text=f"✓ Will receive certificate ({len(will_send)})")
            pane.add(frm_yes, weight=3)
            txt_yes = tk.Text(frm_yes, wrap="none", height=12, font=("Courier", 9))
            sb_yes  = ttk.Scrollbar(frm_yes, command=txt_yes.yview)
            txt_yes.configure(yscrollcommand=sb_yes.set)
            sb_yes.pack(side="right", fill="y"); txt_yes.pack(fill="both", expand=True, padx=4, pady=4)
            for p in sorted(will_send, key=lambda x: x["name"].lower()):
                sessions_cnt = sum(1 for *_, counted in p["session_data"] if counted)
                fuzzy_flag = "  ⚠ name fuzzy-matched" if p.get("fuzzy_matched") else ""
                txt_yes.insert(tk.END,
                    f"{p['name']:<35}  {p['email']:<40}  "
                    f"{fmt_n(p['hours_attended'])}/{fmt_n(p['hours_total'])} hrs  "
                    f"{sessions_cnt} session(s){fuzzy_flag}\n")
            txt_yes.configure(state="disabled")

            # ---- Won't send ----
            frm_no = ttk.LabelFrame(pane, text=f"✗ Will NOT receive certificate ({len(wont_send)})")
            pane.add(frm_no, weight=2)
            txt_no = tk.Text(frm_no, wrap="none", height=8, font=("Courier", 9))
            sb_no  = ttk.Scrollbar(frm_no, command=txt_no.yview)
            txt_no.configure(yscrollcommand=sb_no.set)
            sb_no.pack(side="right", fill="y"); txt_no.pack(fill="both", expand=True, padx=4, pady=4)
            reason = ("No qualifying sessions (no sign-in on any selected date)"
                      if signin_only
                      else "No qualifying sessions (never signed out with certificate = Yes)")
            for p in sorted(wont_send, key=lambda x: x["name"].lower()):
                txt_no.insert(tk.END,
                    f"{p['name']:<35}  {p['email']:<40}  {reason}\n")
            txt_no.configure(state="disabled")

            # ---- Buttons ----
            btn_frm = ttk.Frame(win)
            btn_frm.pack(fill="x", padx=10, pady=8)
            confirmed = [False]

            def _confirm():
                confirmed[0] = True
                win.destroy()

            ttk.Button(btn_frm, text="Confirm — Send Certificates",
                       command=_confirm).pack(side="left")
            ttk.Button(btn_frm, text="Cancel",
                       command=win.destroy).pack(side="left", padx=8)
            ttk.Label(btn_frm, text="Review the lists above before confirming.",
                      foreground="#666").pack(side="left", padx=12)

            win.wait_window()
            if confirmed[0]:
                self._lc_send_all()

        except Exception as e:
            messagebox.showerror("Review", str(e))

    def _lc_send_test(self):
        try:
            event_name = self.lc_training.get().strip() or "Test LC Training"
            trainer    = self.lc_trainer.get().strip()
            location   = self.lc_location.get().strip()
            eila       = self.lc_eila.get().strip()
            if not self.lc_template.get():
                raise ValueError("Select a Certificate Template (DOCX).")
            outdir = self.lc_outdir.get().strip() or os.path.join(os.getcwd(), "lc_output")
            os.makedirs(outdir, exist_ok=True)
            sessions = self._lc_get_sessions() or ([] if self.lc_roster_participants is not None else [("2025-11-12", 3.0)])
            date_range = _lc_date_range(sessions) or self.lc_roster_daterange.get().strip()

            # Roster mode: take the first roster participant with hours.
            test_p = None
            if self.lc_roster_participants is not None:
                for p in self.lc_roster_participants:
                    if p["hours_attended"] > 0:
                        test_p = p
                        break
            else:
                # Load CSVs if not already loaded
                si_df = self.lc_si_df
                so_df = self.lc_so_df
                signin_only = self.lc_use_signin_only.get()
                if si_df is None and self.lc_signin_path.get():
                    si_df = parse_lc_signin_sessions(self.lc_signin_path.get())
                if not signin_only and so_df is None and self.lc_signout_path.get():
                    so_df = parse_lc_signout_sessions(self.lc_signout_path.get())

                # Find first participant with real data who attended >= 2 sessions
                have_sources = si_df is not None and (signin_only or so_df is not None)
                if have_sources:
                    excluded = self._lc_get_excluded()
                    so_for_build = None if signin_only else so_df
                    participants = build_lc_attendance(si_df, so_for_build, sessions, excluded)
                    for p in participants:
                        sessions_attended = sum(1 for _, _, _, _, counted in p["session_data"] if counted)
                        if sessions_attended >= 2:
                            test_p = p
                            break
                    # Fall back to anyone with >= 1 session if nobody has 2
                    if test_p is None:
                        for p in participants:
                            if p["hours_attended"] > 0:
                                test_p = p
                                break

            # Final fallback: synthetic participant with all sessions attended
            if test_p is None:
                hours_total = sum(h for _, h in sessions)
                test_p = {
                    "name":           TEST_NAME,
                    "email":          TEST_EMAIL,
                    "hours_attended": hours_total,
                    "hours_total":    hours_total,
                    "session_data":   [(d, h, True, True, True) for d, h in sessions],
                    "signin_only":    self.lc_use_signin_only.get(),
                }

            # Always send to TEST_EMAIL, never the real participant
            test_p = dict(test_p)
            test_p["email"] = TEST_EMAIL

            ceu_vars = self._lc_get_ceu_vars(sessions)
            fpath = generate_lc_cert(
                test_p["name"], test_p["hours_attended"], test_p["hours_total"],
                event_name, trainer, date_range,
                self.lc_template.get(), outdir,
                location=location, eila=eila, ceu_vars=ceu_vars)
            subj = f"Certificate for {event_name}  {ORG_NAME}"
            body_html = self._lc_build_email_html(test_p, event_name, sessions)
            send_outlook_email(TEST_EMAIL, subj, body_html, fpath, html=True)
            messagebox.showinfo("Test Sent", f"Test LC certificate sent to {TEST_EMAIL}.")
        except Exception as e:
            messagebox.showerror("Test Failed", str(e))

    def _lc_send_all(self):
        try:
            event_name = self.lc_training.get().strip()
            if not event_name:
                raise ValueError("Enter Training Name.")
            trainer  = self.lc_trainer.get().strip()
            location = self.lc_location.get().strip()
            eila     = self.lc_eila.get().strip()
            if not self.lc_template.get():
                raise ValueError("Select a Certificate Template (DOCX).")
            outdir = self.lc_outdir.get().strip() or os.path.join(os.getcwd(), "lc_output")
            os.makedirs(outdir, exist_ok=True)
            sessions = self._lc_get_sessions()
            if not sessions and self.lc_roster_participants is None:
                raise ValueError("No sessions selected.  Click 'Detect Sessions' and check at least one, or load a Precomputed Roster.")
            excluded = self._lc_get_excluded()

            participants = self._lc_get_participants(sessions, excluded)
            total      = len(participants)
            date_range = _lc_date_range(sessions) or self.lc_roster_daterange.get().strip()

            self.lc_progress["value"]   = 0
            self.lc_progress["maximum"] = max(1, total)
            self.lc_prog_label.config(text=f"Sending 0 of {total}…")
            self.update_idletasks()

            issued = []; fail_reasons = []
            start_time = time.time()

            self._lc_cancel = False
            for i, p in enumerate(participants, start=1):
                if self._lc_cancel:
                    self.lc_prog_label.config(text="Send cancelled.")
                    break
                name  = p["name"]
                email = p["email"]
                # Skip participants with no qualifying hours (never signed out with cert=Yes)
                if p["hours_attended"] == 0:
                    fail_reasons.append((
                        name, email,
                        "No qualifying sessions — never completed sign-out with "
                        "certificate request, or declined certificate for all sessions attended"))
                    self.lc_progress["value"] = i
                    self.lc_prog_label.config(text=f"Sending {i} of {total}…  ETA ~ --:--")
                    self.update_idletasks()
                    continue
                try:
                    ceu_vars = self._lc_get_ceu_vars(sessions)
                    fpath = generate_lc_cert(
                        name, p["hours_attended"], p["hours_total"],
                        event_name, trainer, date_range,
                        self.lc_template.get(), outdir,
                        location=location, eila=eila, ceu_vars=ceu_vars)
                    subj      = f"Certificate for {event_name}  {ORG_NAME}"
                    body_html = self._lc_build_email_html(p, event_name, sessions)
                    send_outlook_email(email, subj, body_html, fpath, html=True)
                    issued.append((name, email, p["hours_attended"], p["hours_total"]))
                except Exception as e:
                    fail_reasons.append((name, email, str(e)))

                elapsed   = time.time() - start_time
                avg       = elapsed / max(1, i)
                remaining = max(0, total - i)
                eta       = int(avg * remaining)
                eta_str   = f"{eta//60:02d}:{eta%60:02d}"
                self.lc_progress["value"] = i
                self.lc_prog_label.config(text=f"Sending {i} of {total}…  ETA ~ {eta_str}")
                self.update_idletasks()

            # Completion report
            safe_ev     = safe_filename(event_name)
            report_csv  = os.path.join(outdir, f"LC Certificate Report - {safe_ev}.csv")
            self._lc_send_report(event_name, date_range, sessions, issued, fail_reasons, report_csv)

            messagebox.showinfo("Done", "All LC certificates processed and the report email was sent.")
            self.lc_prog_label.config(text="Done.")
            self._cheer()
        except Exception as e:
            messagebox.showerror("Send All Certificates", str(e))

    def _lc_send_report(self, event_name, date_range, sessions, issued, fails, report_csv_path):
        try:
            report_dir = os.path.dirname(report_csv_path)
            if report_dir: os.makedirs(report_dir, exist_ok=True)
            fmt_n = lambda n: f"{float(n):g}"
            with open(report_csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["Name", "Email", "Hours Attended", "Hours Total"])
                for n, e, h_att, h_tot in sorted(issued, key=lambda x: x[0].lower()):
                    w.writerow([n, e, h_att, h_tot])
                w.writerow([])
                w.writerow(["Not Issued", "Email", "Reason"])
                for n, e, r in fails:
                    w.writerow([n, e, r])
            if not win32: raise RuntimeError("Outlook not available.")
            hours_total = sum(h for _, h in sessions)
            lines = [
                f"LC Certificate Report: {event_name}",
                f"Date Range: {date_range}",
                f"Total available hours: {fmt_n(hours_total)}",
                "",
                f"Certificates Issued ({len(issued)})",
            ]
            for n, e, h_att, h_tot in sorted(issued, key=lambda x: x[0].lower()):
                lines.append(f"  {n} <{e}> — {fmt_n(h_att)} / {fmt_n(h_tot)} hrs")
            lines.append("")
            if fails:
                lines.append(f"Not Issued ({len(fails)})")
                for n, e, r in fails:
                    lines.append(f"  {n} <{e}> — {r}")
            else:
                lines.append("Not Issued: None")
            subj = f"LC Certificate Report: {event_name}"
            send_outlook_email(TEST_EMAIL, subj, md_to_html("\n".join(lines)), report_csv_path, html=True)
        except Exception as e:
            messagebox.showwarning("Report Email", f"Report email failed: {e}")

    def _lc_start_over(self):
        for w in (self.lc_training, self.lc_trainer, self.lc_location, self.lc_eila,
                  self.lc_signin_path, self.lc_signout_path, self.lc_template, self.lc_outdir,
                  self.lc_roster_path, self.lc_roster_daterange):
            try: w.delete(0, tk.END)
            except Exception: pass
        self.lc_si_df = None
        self.lc_so_df = None
        self.lc_roster_participants = None
        self.lc_use_signin_only.set(False)
        self._lc_cancel = False
        self._lc_clear_session_rows()
        self._lc_no_sessions_msg()
        self.lc_msg.delete("1.0", tk.END)
        self.lc_msg.insert("1.0", DEFAULT_LC_EMAIL)
        self.lc_progress["value"] = 0
        self.lc_prog_label.config(text="Idle.")
        # Reset CEU fields
        for v in (self.lc_ceu_sw, self.lc_ceu_psych, self.lc_ceu_eila,
                  self.lc_ceu_frsky, self.lc_ceu_aswb, self.lc_ceu_lpcc):
            v.set(False)
        if hasattr(self, 'lc_frsky'):
            self.lc_frsky.delete(0, tk.END)
        if hasattr(self, 'lc_aswb_period'):
            self.lc_aswb_period.delete(0, tk.END)
        self._lc_ceu_toggle()

    # ---- LC log-to-master helpers ----

    def _lc_log_refresh_sessions(self):
        """Refresh the session picker dropdown after Detect Sessions runs."""
        sessions = self._lc_get_sessions()   # list of (date_str, hours)
        if not sessions:
            self.lc_log_session_cb["values"] = []
            self.lc_log_session_var.set("(no sessions detected)")
            return
        labels = [f"{d}  ({h} hrs)" for d, h in sessions]
        self.lc_log_session_cb["values"] = labels
        self.lc_log_session_cb.set(labels[0])

    def _lc_log_prefill(self):
        """Prefill log fields from the currently selected session."""
        idx = self.lc_log_session_cb.current()
        sessions = self._lc_get_sessions()
        if idx < 0 or not sessions or idx >= len(sessions):
            messagebox.showwarning("No Session",
                "Please detect sessions first, then pick one from the dropdown.")
            return
        date_str, hours = sessions[idx]
        self._lc_log_fill_for_date(date_str, hours)

    def _lc_log_fill_for_date(self, date_str: str, hours):
        """Populate all log fields for a given session date."""
        def _set(key, val):
            w = self.lc_log_entries.get(key)
            if w is None: return
            if isinstance(w, tk.Text):
                w.delete("1.0", tk.END); w.insert("1.0", val)
            else:
                w.delete(0, tk.END); w.insert(0, str(val))

        # Session number = position in sorted session list
        sessions = self._lc_get_sessions()
        dates = [d for d, _ in sessions]
        session_num = (dates.index(date_str) + 1) if date_str in dates else ""
        session_label = f"Learning Session {session_num}" if session_num else date_str

        lc_name = self.lc_training.get().strip() or "Learning Collaborative"
        training_name = f"{lc_name}: {session_label}"

        # Attendance from sign-in count on this date
        si_count = 0
        psych_count = 0
        si_path = self.lc_signin_path.get().strip() if hasattr(self, "lc_signin_path") else ""
        if hasattr(self, "lc_si_df") and self.lc_si_df is not None:
            day_df = self.lc_si_df[self.lc_si_df["DATE"] == date_str]
            si_count = len(day_df)
        # Psych count — re-read raw sign-in CSV filtered to this date
        if si_path and os.path.isfile(si_path):
            psych_count = _lc_count_psych_for_date(si_path, date_str)

        # CEU string from LC CEU checkboxes
        ceu_parts = []
        if self.lc_ceu_sw.get():    ceu_parts.append("SW")
        if self.lc_ceu_psych.get(): ceu_parts.append("Psy")
        if self.lc_ceu_lpcc.get():  ceu_parts.append("LPCC")
        if self.lc_ceu_aswb.get():  ceu_parts.append("ASWB")
        if self.lc_ceu_eila.get():  ceu_parts.append("EILA")
        if self.lc_ceu_frsky.get(): ceu_parts.append("FRSKY")
        ceu_str = ", ".join(ceu_parts) if ceu_parts else "none"

        # Location info
        location = self.lc_location.get().strip() if hasattr(self, "lc_location") else ""
        city = location.split(",")[0].strip() if "," in location else location

        _set("Training",        training_name)
        _set("Date",            date_str)
        _set("Training Length", str(hours))
        _set("Attendance",      str(si_count) if si_count else "")
        _set("Psych Attendance",str(psych_count))
        _set("CEU",             ceu_str)
        _set("City",            city)
        _set("Venue",           self.lc_location.get().strip()
                                if hasattr(self, "lc_location") else "Zoom")

    def _lc_log_get_row(self) -> list:
        """Read lc_log_entries in column order and return a list of values."""
        headers = [
            "Date","Program","Training","Trainers","Training Length","Attendance",
            "Psych Attendance","Target Audience","Evals/Outcome","CEU","City","County","Venue"
        ]
        row = []
        for h in headers:
            w = self.lc_log_entries.get(h)
            if w is None:
                row.append("")
            elif isinstance(w, tk.Text):
                row.append(w.get("1.0", tk.END).strip())
            else:
                row.append(w.get().strip())
        # Trainers comes from main LC Event Details trainer field
        if not row[3]:   # index 3 = Trainers
            row[3] = self.lc_trainer.get().strip() if hasattr(self, "lc_trainer") else ""
        return row

    def _lc_log_session(self):
        """Log the currently filled session to Training Information tab."""
        try:
            path = self.mtl_path.get().strip()
            if not path:
                raise ValueError("Master Training List path is blank.")
            directory = os.path.dirname(path)
            if directory and not os.path.isdir(directory):
                raise ValueError(
                    "Path not reachable:\n" + path +
                    "\n\nCheck your network connection.")
            row_vals = self._lc_log_get_row()
            if not row_vals[0]:   # Date
                raise ValueError("Date is blank — prefill from a session first.")
            if not row_vals[2]:   # Training
                raise ValueError("Training Name is blank.")
            append_row_openpyxl(path, "Training Information", row_vals)
            name = row_vals[2]
            self.lc_log_status.config(
                text=f"✔ Logged: {name[:50]}")
            messagebox.showinfo("Master Training List",
                f"Session logged successfully:\n{name}")
        except Exception as e:
            messagebox.showerror("Log Session", str(e))

    def _lc_log_all_sessions(self):
        """Log every detected session to Training Information in one pass."""
        sessions = self._lc_get_sessions()
        if not sessions:
            messagebox.showwarning("No Sessions",
                "No sessions detected. Run 'Detect Sessions' first.")
            return
        path = self.mtl_path.get().strip()
        if not path:
            messagebox.showerror("Log All", "Master Training List path is blank.")
            return
        directory = os.path.dirname(path)
        if directory and not os.path.isdir(directory):
            messagebox.showerror("Log All",
                f"Path not reachable:\n{path}\n\nCheck your network connection.")
            return
        if not messagebox.askyesno("Log All Sessions",
                f"This will add {len(sessions)} rows to the Training Information tab.\n"
                "Continue?"):
            return
        logged = 0
        errors = []
        for date_str, hours in sessions:
            try:
                self._lc_log_fill_for_date(date_str, hours)
                row_vals = self._lc_log_get_row()
                append_row_openpyxl(path, "Training Information", row_vals)
                logged += 1
            except Exception as exc:
                errors.append(f"{date_str}: {exc}")
        msg = f"Logged {logged} of {len(sessions)} sessions."
        if errors:
            msg += "\n\nErrors:\n" + "\n".join(errors)
        self.lc_log_status.config(text=f"✔ Logged {logged} sessions.")
        messagebox.showinfo("Log All Sessions", msg)


    # ================================================================
    # Evaluation Report tab
    # ================================================================

    def _build_eval_report_tab(self, root):
        if not REPORTLAB_OK:
            ttk.Label(root,
                      text="reportlab is not installed. Run:  pip install reportlab",
                      foreground="red").pack(padx=20, pady=20)
            return

        # ---- Training Details Source ----
        frm_src = ttk.LabelFrame(root, text="Training Details Source")
        frm_src.pack(fill="x", padx=12, pady=(10, 6))
        self._eval_source_var = tk.IntVar(value=1)
        ttk.Radiobutton(frm_src, text="Pull from Standard Training tab",
                        variable=self._eval_source_var, value=1,
                        command=self._eval_source_toggle).grid(row=0, column=0, sticky="w", padx=10, pady=6)
        ttk.Radiobutton(frm_src, text="Pull from Learning Collaborative tab",
                        variable=self._eval_source_var, value=2,
                        command=self._eval_source_toggle).grid(row=0, column=1, sticky="w", padx=10, pady=6)
        ttk.Radiobutton(frm_src, text="Enter manually",
                        variable=self._eval_source_var, value=3,
                        command=self._eval_source_toggle).grid(row=0, column=2, sticky="w", padx=10, pady=6)
        self._eval_manual_frm = ttk.Frame(frm_src)
        self._eval_manual_frm.grid(row=1, column=0, columnspan=3, sticky="ew", padx=8, pady=(0, 8))
        ttk.Label(self._eval_manual_frm, text="Training Name:", width=14).grid(row=0, column=0, sticky="w", padx=4, pady=3)
        self.eval_manual_name = ttk.Entry(self._eval_manual_frm, width=52)
        self.eval_manual_name.grid(row=0, column=1, sticky="w", padx=4, pady=3)
        ttk.Label(self._eval_manual_frm, text="Trainers:", width=14).grid(row=1, column=0, sticky="w", padx=4, pady=3)
        self.eval_manual_trainers = ttk.Entry(self._eval_manual_frm, width=52)
        self.eval_manual_trainers.grid(row=1, column=1, sticky="w", padx=4, pady=3)
        ttk.Label(self._eval_manual_frm, text="Date(s):", width=14).grid(row=2, column=0, sticky="w", padx=4, pady=3)
        self.eval_manual_date = ttk.Entry(self._eval_manual_frm, width=52)
        self.eval_manual_date.grid(row=2, column=1, sticky="w", padx=4, pady=3)
        self._eval_source_toggle()

        # ---- Evaluation File ----
        frm1 = ttk.LabelFrame(root, text="Evaluation File")
        frm1.pack(fill="x", padx=12, pady=6)
        self.eval_path = self._eval_file_row(
            frm1, "CSV or XLSX:", 0,
            filetypes=[("CSV or Excel", "*.csv *.xlsx"), ("CSV", "*.csv"), ("Excel", "*.xlsx")],
            on_pick=self._eval_on_pick)
        ttk.Button(frm1, text="Load & Detect Sessions",
                   command=self._eval_detect_sessions).grid(row=1, column=1, sticky="w", padx=8, pady=(4, 8))

        # ---- Sessions ----
        frm2 = ttk.LabelFrame(root, text="Sessions  (uncheck any test dates — edit labels to match session names)")
        frm2.pack(fill="x", padx=12, pady=6)
        ttk.Label(frm2, text="Include", width=8).grid(row=0, column=0, padx=(10, 2), pady=4)
        ttk.Label(frm2, text="Session Label (editable)").grid(row=0, column=1, padx=(2, 4), pady=4, sticky="w")
        ttk.Label(frm2, text="Responses").grid(row=0, column=2, padx=(4, 8), pady=4, sticky="w")
        self._eval_sessions_frame = frm2
        self._eval_no_sessions_msg()

        # ---- Training Manager Notes (goes in email, not PDF) ----
        frm3 = ttk.LabelFrame(root, text="Training Manager Notes  (optional — included in the email, not the PDF)")
        frm3.pack(fill="both", expand=False, padx=12, pady=6)
        self.eval_notes = tk.Text(frm3, height=4, wrap="word")
        self.eval_notes.pack(fill="both", expand=True, padx=8, pady=8)

        # ---- Output Folder (required) ----
        frm4 = ttk.LabelFrame(root, text="Output Folder  (required — report PDF and source CSV will be saved here)")
        frm4.pack(fill="x", padx=12, pady=6)
        self.eval_outdir = self._eval_folder_row(frm4, "Save to:", 0)

        # ---- Recipients ----
        frm5 = ttk.LabelFrame(root, text="Send Report To")
        frm5.pack(fill="x", padx=12, pady=6)
        people = [
            ("Ginny Sprang",      "sprang@uky.edu"),
            ("Tracy Clemans",     "tracy.clemans@uky.edu"),
            ("Leah Riggs",        "leah.riggs@uky.edu"),
            ("Stephanie Gusler",  "stephanie.gusler@uky.edu"),
            ("Alex Clark",        "alex-clark@uky.edu"),
            ("Jessica Eslinger",  "j.g.eslinger@uky.edu"),
            ("Holly Huber",       "hhubergifford@uky.edu"),
            ("Adrienne Whitt",    "adrienne.whitt-woosley@uky.edu"),
            ("Josh Fisherkeller", "jafish0@uky.edu"),
        ]
        self._eval_checks = []
        rec_frame = ttk.Frame(frm5)
        rec_frame.pack(fill="x", padx=8, pady=6)
        for i, (name, email) in enumerate(people):
            var = tk.BooleanVar(value=(email in ("sprang@uky.edu", "jafish0@uky.edu")))
            ttk.Checkbutton(rec_frame, text=name, variable=var).grid(
                row=i // 3, column=i % 3, sticky="w", padx=10, pady=2)
            self._eval_checks.append((var, name, email))

        # ---- Buttons ----
        btn_frm = ttk.Frame(root)
        btn_frm.pack(fill="x", padx=12, pady=8)
        ttk.Button(btn_frm, text="Generate Report (PDF)",
                   command=lambda: self._eval_generate_report(send=False)).pack(side="left")
        ttk.Button(btn_frm, text="Generate & Send",
                   command=lambda: self._eval_generate_report(send=True)).pack(side="left", padx=(8, 0))
        ttk.Button(btn_frm, text="Start Over",
                   command=self._eval_start_over).pack(side="left", padx=8)

        # ---- Status ----
        frm6 = ttk.LabelFrame(root, text="Status")
        frm6.pack(fill="x", padx=12, pady=(6, 12))
        self.eval_status = ttk.Label(frm6, text="Idle.")
        self.eval_status.pack(anchor="w", padx=12, pady=8)

    def _eval_source_toggle(self):
        src = self._eval_source_var.get()
        if src == 3:
            self._eval_manual_frm.grid()
        else:
            self._eval_manual_frm.grid_remove()

    def _eval_get_event_info(self):
        """Return (training_name, trainers, date) based on the source selector."""
        src = self._eval_source_var.get()
        if src == 1:
            ev = self.get_event()
            return ev.get("TRAINING", ""), ev.get("TRAINER", ""), ev.get("DATE", "")
        elif src == 2:
            name    = self.lc_training.get().strip() if hasattr(self, "lc_training") else ""
            trainer = self.lc_trainer.get().strip()  if hasattr(self, "lc_trainer")  else ""
            return name, trainer, ""
        else:
            return (self.eval_manual_name.get().strip(),
                    self.eval_manual_trainers.get().strip(),
                    self.eval_manual_date.get().strip())

    def _eval_file_row(self, parent, label, row, filetypes=None, on_pick=None):
        if filetypes is None:
            filetypes = [("CSV", "*.csv")]
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8, 6), pady=4)
        v = ttk.Entry(parent, width=56)
        v.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        def browse():
            p = filedialog.askopenfilename(title=label, filetypes=filetypes)
            if p:
                v.delete(0, tk.END); v.insert(0, p)
                if on_pick: on_pick(p)
        ttk.Button(parent, text="Browse\u2026", command=browse).grid(row=row, column=2, padx=6)
        return v

    def _eval_folder_row(self, parent, label, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(8, 6), pady=4)
        v = ttk.Entry(parent, width=56)
        v.grid(row=row, column=1, sticky="w", padx=8, pady=4)
        def browse():
            p = filedialog.askdirectory(title=label)
            if p:
                v.delete(0, tk.END); v.insert(0, p)
        ttk.Button(parent, text="Browse\u2026", command=browse).grid(row=row, column=2, padx=6)
        return v

    def _eval_on_pick(self, path):
        self._eval_df = None
        try:
            _, sessions = parse_eval_file(path)
            self._eval_df = sessions
        except Exception as e:
            messagebox.showerror("Evaluation File", str(e))

    def _eval_no_sessions_msg(self):
        ttk.Label(self._eval_sessions_frame,
                  text="No sessions detected yet.  Upload evaluation file then click 'Load & Detect Sessions'.",
                  foreground="gray").grid(row=1, column=0, columnspan=4, padx=10, pady=6, sticky="w")

    def _eval_clear_session_rows(self):
        frm = self._eval_sessions_frame
        for widget in list(frm.winfo_children()):
            try:
                info = widget.grid_info()
                if info and int(info.get("row", 0)) >= 1:
                    widget.destroy()
            except Exception:
                pass
        self._eval_session_rows = []

    def _eval_detect_sessions(self):
        try:
            path = self.eval_path.get().strip()
            if not path:
                raise ValueError("Please select an evaluation file first.")
            _, sessions = parse_eval_file(path)
            self._eval_df = sessions
            self._eval_clear_session_rows()
            if not sessions:
                self._eval_no_sessions_msg()
                return
            frm = self._eval_sessions_frame
            for i, date_str in enumerate(sorted(sessions.keys())):
                df = sessions[date_str]
                n = len(df)
                likely_real = n >= 3
                var = tk.BooleanVar(value=likely_real)
                ttk.Checkbutton(frm, variable=var).grid(row=i+1, column=0, padx=(14, 2), pady=2)
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    fmt = "%#m/%#d/%Y" if os.name == "nt" else "%-m/%-d/%Y"
                    date_display = dt.strftime(fmt)
                except Exception:
                    date_display = date_str
                label_e = ttk.Entry(frm, width=32)
                label_e.insert(0, date_display)
                label_e.grid(row=i+1, column=1, padx=(2, 4), pady=2, sticky="w")
                ttk.Label(frm, text=f"{n} response{'s' if n != 1 else ''}", width=14).grid(
                    row=i+1, column=2, padx=(4, 8), pady=2, sticky="w")
                if not likely_real:
                    ttk.Label(frm, text="\u2190 likely test", foreground="#888").grid(
                        row=i+1, column=3, padx=4, pady=2, sticky="w")
                self._eval_session_rows.append((var, date_str, label_e))
            real = sum(1 for v, _, _ in self._eval_session_rows if v.get())
            messagebox.showinfo("Sessions Detected",
                f"Found {len(sessions)} session date(s).  {real} appear to have real responses.\n\n"
                "Edit the label in each row to match your session name "
                "(e.g. 'Learning Session 1 - 11/12/25').")
        except Exception as e:
            messagebox.showerror("Detect Sessions", str(e))

    def _eval_generate_report(self, send=False):
        import shutil
        try:
            if not REPORTLAB_OK:
                raise RuntimeError("reportlab is not installed.  Run: pip install reportlab")
            src_path = self.eval_path.get().strip()
            if not src_path:
                raise ValueError("Please select an evaluation file.")
            if not self._eval_session_rows:
                raise ValueError("No sessions detected.  Click 'Load & Detect Sessions' first.")
            selected = [(label_e.get().strip(), date_str)
                        for (var, date_str, label_e) in self._eval_session_rows if var.get()]
            if not selected:
                raise ValueError("No sessions are checked.  Please select at least one session.")
            outdir = self.eval_outdir.get().strip()
            if not outdir:
                raise ValueError("Please select an output folder.")
            if not os.path.isdir(outdir):
                raise ValueError(f"Output folder does not exist:\n{outdir}\n\nPlease browse to an existing folder.")
            _, all_sessions = parse_eval_file(src_path)
            sessions_data = [(label, all_sessions[ds]) for label, ds in selected if ds in all_sessions]
            training_name, trainers, date_str = self._eval_get_event_info()
            training_name = training_name or "Training Evaluation Report"
            notes = self.eval_notes.get("1.0", tk.END).strip()
            out_pdf = os.path.join(outdir, f"Evaluation Report - {safe_filename(training_name)}.pdf")
            self.eval_status.config(text="Generating PDF\u2026")
            self.update_idletasks()
            # Generate PDF (notes go in email, not PDF)
            generate_eval_pdf(training_name, sessions_data, out_pdf)
            # Copy source CSV/XLSX to output folder
            dst_csv = os.path.join(outdir, os.path.basename(src_path))
            if os.path.abspath(src_path) != os.path.abspath(dst_csv):
                shutil.copy2(src_path, dst_csv)
            self.eval_status.config(text=f"Saved: {os.path.basename(out_pdf)}  |  {os.path.basename(dst_csv)}")
            self._cheer()
            if send:
                self._eval_do_send(training_name, trainers, date_str, notes, out_pdf)
            else:
                messagebox.showinfo("Report Generated",
                    f"Saved to:\n  {os.path.basename(out_pdf)}\n  {os.path.basename(dst_csv)}\n\nFolder:\n  {outdir}")
        except Exception as e:
            self.eval_status.config(text="Error \u2014 see dialog.")
            messagebox.showerror("Evaluation Report", str(e))

    def _eval_do_send(self, training_name, trainers, date_str, notes, pdf_path):
        try:
            if not win32:
                raise RuntimeError("Outlook not available.")
            selected = [(n, e) for (v, n, e) in self._eval_checks if v.get()]
            if not selected:
                messagebox.showerror("Recipients", "Select at least one recipient.")
                return
            subj = f"Evaluation Report: {training_name}".strip()
            if date_str:
                subj += f" {date_str}"
            body_lines = [f"Hello,\n\nPlease find attached the evaluation report for {training_name}."]
            if trainers:
                body_lines.append(f"\nTrainers: {trainers}")
            if date_str:
                body_lines.append(f"Date(s): {date_str}")
            if notes:
                body_lines.append(f"\nTraining Manager Notes:\n{notes}")
            body_lines.append(
                "\n\nJoshua Fisherkeller, MSW\n"
                "Education and Training Manager\n"
                "University of Kentucky Center on Trauma and Children\n"
                "(859) 218-6941"
            )
            body = "\n".join(body_lines)
            for name, email in selected:
                send_outlook_email(email, subj, md_to_html(body), pdf_path, html=True)
            self.eval_status.config(text=f"Report sent to {len(selected)} recipient(s).")
            messagebox.showinfo("Report Sent",
                "Evaluation report sent to:\n" + "\n".join(f"  {n} <{e}>" for n, e in selected))
        except Exception as e:
            messagebox.showerror("Send Report", str(e))

    def _eval_start_over(self):
        if hasattr(self, "eval_path"):
            self.eval_path.delete(0, tk.END)
        self._eval_df = None
        self._eval_clear_session_rows()
        self._eval_no_sessions_msg()
        if hasattr(self, "eval_notes"):
            self.eval_notes.delete("1.0", tk.END)
        if hasattr(self, "eval_outdir"):
            self.eval_outdir.delete(0, tk.END)
        if hasattr(self, "eval_manual_name"):
            self.eval_manual_name.delete(0, tk.END)
            self.eval_manual_trainers.delete(0, tk.END)
            self.eval_manual_date.delete(0, tk.END)
        if hasattr(self, "eval_status"):
            self.eval_status.config(text="Idle.")

    # ================================================================
    # Training Request tab
    # ================================================================

    TRAINER_CHOICES_REQ = [
        "Tracy Clemans", "Leah Riggs", "Josh Fisherkeller", "No Preference"
    ]

    def _build_request_tab(self, root):
        if not REPORTLAB_OK:
            ttk.Label(root,
                      text="reportlab is not installed.  pip install reportlab",
                      foreground="red").pack(padx=20, pady=20)
            return

        # ---- Mode selector ----
        frm_mode = ttk.LabelFrame(root, text="Mode")
        frm_mode.pack(fill="x", padx=12, pady=(10, 4))

        self._req_mode = tk.IntVar(value=1)   # 1 = New, 2 = Load Pending
        ttk.Radiobutton(frm_mode, text="New Request",
                        variable=self._req_mode, value=1,
                        command=self._req_mode_toggle).grid(
            row=0, column=0, sticky="w", padx=10, pady=6)
        ttk.Radiobutton(frm_mode, text="Load Pending Request:",
                        variable=self._req_mode, value=2,
                        command=self._req_mode_toggle).grid(
            row=0, column=1, sticky="w", padx=(14, 4), pady=6)
        self._req_pending_var = tk.StringVar()
        self._req_pending_cb = ttk.Combobox(frm_mode, textvariable=self._req_pending_var,
                                             state="disabled", width=48)
        self._req_pending_cb.grid(row=0, column=2, sticky="w", padx=4, pady=6)
        self._req_pending_cb.bind("<<ComboboxSelected>>", self._req_on_pending_select)
        ttk.Button(frm_mode, text="↺ Refresh",
                   command=self._req_refresh_pending_list).grid(
            row=0, column=3, padx=8, pady=6)

        # ---- Request Details ----
        frm_det = ttk.LabelFrame(root, text="Request Details")
        frm_det.pack(fill="x", padx=12, pady=4)
        frm_det.columnconfigure(1, weight=1)

        def lbl(txt, r, c=0, span=1, bold=False):
            f = tk.font.Font(weight="bold") if bold else None
            kw = {"font": f} if bold else {}
            ttk.Label(frm_det, text=txt, **kw).grid(
                row=r, column=c, sticky="w", padx=(8, 4), pady=3)

        def ent(r, c=1, width=54):
            e = ttk.Entry(frm_det, width=width)
            e.grid(row=r, column=c, sticky="w", padx=8, pady=3)
            return e

        lbl("Request Date:",       0); self.req_date       = ent(0)
        lbl("Organization/Who:",   1); self.req_org        = ent(1)
        lbl("Training Requested:", 2); self.req_training   = ent(2)
        lbl("Requested Date:",     3); self.req_req_date   = ent(3)
        lbl("Location/Where:",     4); self.req_location   = ent(4)

        # Default today
        self.req_date.insert(0, datetime.now().strftime("%m/%d/%Y"))

        # ---- Participant Info ----
        frm_part = ttk.LabelFrame(root, text="Participant Information")
        frm_part.pack(fill="x", padx=12, pady=4)
        frm_part.columnconfigure(1, weight=1)

        ttk.Label(frm_part, text="# of Participants:").grid(
            row=0, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_num_part = ttk.Entry(frm_part, width=20)
        self.req_num_part.grid(row=0, column=1, sticky="w", padx=8, pady=3)

        ttk.Label(frm_part, text="Participant Roles:").grid(
            row=1, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_roles = ttk.Entry(frm_part, width=54)
        self.req_roles.grid(row=1, column=1, sticky="w", padx=8, pady=3)

        # ---- Logistics ----
        frm_log = ttk.LabelFrame(root, text="Logistics")
        frm_log.pack(fill="x", padx=12, pady=4)
        frm_log.columnconfigure(1, weight=1)

        ttk.Label(frm_log, text="Preferred Trainer:").grid(
            row=0, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_trainer = ttk.Combobox(frm_log, values=self.TRAINER_CHOICES_REQ,
                                         width=30, state="normal")
        self.req_trainer.grid(row=0, column=1, sticky="w", padx=8, pady=3)

        ttk.Label(frm_log, text="Mileage Cost:").grid(
            row=1, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_mileage = ttk.Entry(frm_log, width=20)
        self.req_mileage.grid(row=1, column=1, sticky="w", padx=8, pady=3)

        ttk.Label(frm_log, text="Other Travel (Hotel, Parking, etc.):").grid(
            row=2, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_travel = ttk.Entry(frm_log, width=40)
        self.req_travel.grid(row=2, column=1, sticky="w", padx=8, pady=3)

        # ---- Billing Contact ----
        frm_bill = ttk.LabelFrame(root, text="Billing Contact")
        frm_bill.pack(fill="x", padx=12, pady=4)
        frm_bill.columnconfigure(1, weight=1)

        ttk.Label(frm_bill, text="Contact Name:").grid(
            row=0, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_bill_name = ttk.Entry(frm_bill, width=40)
        self.req_bill_name.grid(row=0, column=1, sticky="w", padx=8, pady=3)

        ttk.Label(frm_bill, text="Contact Email:").grid(
            row=1, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_bill_email = ttk.Entry(frm_bill, width=40)
        self.req_bill_email.grid(row=1, column=1, sticky="w", padx=8, pady=3)

        ttk.Label(frm_bill, text="Contact Phone:").grid(
            row=2, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_bill_phone = ttk.Entry(frm_bill, width=30)
        self.req_bill_phone.grid(row=2, column=1, sticky="w", padx=8, pady=3)

        # ---- Additional Comments ----
        frm_notes = ttk.LabelFrame(root, text="Additional Comments")
        frm_notes.pack(fill="x", padx=12, pady=4)
        self.req_comments = tk.Text(frm_notes, height=3, width=70, wrap="word")
        self.req_comments.pack(fill="x", padx=8, pady=6)

        # ---- Submit buttons ----
        frm_sub = ttk.Frame(root)
        frm_sub.pack(fill="x", padx=12, pady=6)
        ttk.Button(frm_sub, text="📤  Send Request to Dr. Sprang",
                   command=lambda: self._req_send_to_sprang(test=False)).pack(
            side="left", padx=4)
        ttk.Button(frm_sub, text="🧪  Send Test to Myself",
                   command=lambda: self._req_send_to_sprang(test=True)).pack(
            side="left", padx=4)
        ttk.Button(frm_sub, text="Start Over",
                   command=self._req_start_over).pack(side="left", padx=4)
        self.req_status = ttk.Label(frm_sub, text="", foreground="green")
        self.req_status.pack(side="left", padx=12)

        # ==== QUOTE SECTION ====
        ttk.Separator(root, orient="horizontal").pack(fill="x", padx=12, pady=8)
        ttk.Label(root, text="QUOTE  —  complete after Dr. Sprang provides pricing",
                  font=("TkDefaultFont", 10, "bold")).pack(padx=12, anchor="w")

        # Quote line items (up to 4 rows)
        frm_qi = ttk.LabelFrame(root, text="Quote Line Items")
        frm_qi.pack(fill="x", padx=12, pady=6)

        ttk.Label(frm_qi, text="Training Event / Description",
                  font=("TkDefaultFont", 9, "bold")).grid(
            row=0, column=0, padx=8, pady=4, sticky="w")
        ttk.Label(frm_qi, text="Date",
                  font=("TkDefaultFont", 9, "bold")).grid(
            row=0, column=1, padx=8, pady=4)
        ttk.Label(frm_qi, text="Cost ($)",
                  font=("TkDefaultFont", 9, "bold")).grid(
            row=0, column=2, padx=8, pady=4)

        self._req_line_items = []   # list of (desc_entry, date_entry, cost_entry)
        vcmd = (root.register(self._req_quote_changed), "%P")
        for i in range(4):
            d = ttk.Entry(frm_qi, width=40)
            d.grid(row=i+1, column=0, padx=8, pady=2, sticky="w")
            dt = ttk.Entry(frm_qi, width=12)
            dt.grid(row=i+1, column=1, padx=8, pady=2)
            c = ttk.Entry(frm_qi, width=12,
                          validate="key", validatecommand=vcmd)
            c.grid(row=i+1, column=2, padx=8, pady=2)
            self._req_line_items.append((d, dt, c))

        ttk.Label(frm_qi, text="Total:").grid(
            row=5, column=1, sticky="e", padx=8, pady=6)
        self.req_quote_total = ttk.Label(frm_qi, text="$0.00",
                                          font=("TkDefaultFont", 9, "bold"))
        self.req_quote_total.grid(row=5, column=2, sticky="w", padx=8, pady=6)

        # Quote email section
        frm_qe = ttk.LabelFrame(root, text="Quote Email")
        frm_qe.pack(fill="x", padx=12, pady=4)
        frm_qe.columnconfigure(1, weight=1)

        ttk.Label(frm_qe, text="Send To:").grid(
            row=0, column=0, sticky="w", padx=(8,4), pady=3)
        self.req_quote_to = ttk.Entry(frm_qe, width=46)
        self.req_quote_to.grid(row=0, column=1, sticky="w", padx=8, pady=3)

        ttk.Label(frm_qe, text="Email Body:").grid(
            row=1, column=0, sticky="nw", padx=(8,4), pady=3)
        self.req_quote_body = tk.Text(frm_qe, height=10, width=70, wrap="word")
        self.req_quote_body.grid(row=1, column=1, sticky="ew", padx=8, pady=4)

        frm_qbtns = ttk.Frame(root)
        frm_qbtns.pack(fill="x", padx=12, pady=4)
        ttk.Button(frm_qbtns, text="✉  Preview & Send Quote",
                   command=self._req_preview_send_quote).pack(side="left", padx=4)
        ttk.Button(frm_qbtns, text="✔  Mark as Accepted",
                   command=lambda: self._req_mark_resolved("accepted")).pack(
            side="left", padx=4)
        ttk.Button(frm_qbtns, text="✘  Mark as Declined",
                   command=lambda: self._req_mark_resolved("declined")).pack(
            side="left", padx=4)
        ttk.Button(frm_qbtns, text="Mark as Cancelled",
                   command=lambda: self._req_mark_resolved("cancelled")).pack(
            side="left", padx=4)

        self.req_quote_status = ttk.Label(frm_qbtns, text="", foreground="green")
        self.req_quote_status.pack(side="left", padx=12)

        # Populate the quote email body with a default on startup
        self._req_refresh_pending_list()
        self._req_populate_quote_email()

    # ---- helpers ----

    def _req_mode_toggle(self):
        if self._req_mode.get() == 2:
            self._req_pending_cb.config(state="readonly")
            self._req_refresh_pending_list()
        else:
            self._req_pending_cb.config(state="disabled")
            self._req_current_id = None

    def _req_refresh_pending_list(self):
        """Reload the Load Pending dropdown from JSON store."""
        records = _req_load_all_json(self._req_store)
        pending = [r for r in records
                   if r.get("status") in ("submitted", "quoted")]
        labels = []
        for r in pending:
            labels.append(f"{r.get('request_date','')}  |  {r.get('org_name','')}  "
                          f"[{r.get('status','')}]")
        self._req_pending_records = pending
        self._req_pending_cb["values"] = labels
        if not labels:
            self._req_pending_cb.set("(no pending requests)")
        elif self._req_pending_cb.get() not in labels:
            self._req_pending_cb.set("")

    def _req_on_pending_select(self, event=None):
        idx = self._req_pending_cb.current()
        if idx < 0 or idx >= len(getattr(self, "_req_pending_records", [])):
            return
        data = self._req_pending_records[idx]
        self._req_populate_form(data)
        self._req_current_id = data.get("id")
        # Pre-fill quote send-to from billing email
        self.req_quote_to.delete(0, tk.END)
        self.req_quote_to.insert(0, data.get("billing_email", ""))
        # Pre-fill quote line items if they exist
        for i, (de, dte, ce) in enumerate(self._req_line_items):
            li = data.get("line_items", [{}]*4)
            row = li[i] if i < len(li) else {}
            de.delete(0, tk.END);  de.insert(0, row.get("desc", ""))
            dte.delete(0, tk.END); dte.insert(0, row.get("date", ""))
            ce.delete(0, tk.END);  ce.insert(0, row.get("cost", ""))
        self._req_quote_changed()
        self._req_populate_quote_email()

    def _req_populate_form(self, data):
        """Fill all form widgets from a data dict."""
        def _set(widget, val):
            if isinstance(widget, tk.Text):
                widget.delete("1.0", tk.END)
                widget.insert("1.0", val or "")
            else:
                widget.delete(0, tk.END)
                widget.insert(0, val or "")
        _set(self.req_date,       data.get("request_date", ""))
        _set(self.req_org,        data.get("org_name", ""))
        _set(self.req_training,   data.get("training_requested", ""))
        _set(self.req_req_date,   data.get("requested_date", ""))
        _set(self.req_location,   data.get("location", ""))
        _set(self.req_num_part,   data.get("num_participants", ""))
        _set(self.req_roles,      data.get("participant_roles", ""))
        self.req_trainer.set(data.get("preferred_trainer", ""))
        _set(self.req_mileage,    data.get("mileage_cost", ""))
        _set(self.req_travel,     data.get("other_travel_cost", ""))
        _set(self.req_bill_name,  data.get("billing_contact", ""))
        _set(self.req_bill_email, data.get("billing_email", ""))
        _set(self.req_bill_phone, data.get("billing_phone", ""))
        _set(self.req_comments,   data.get("additional_comments", ""))

    def _req_get_form_data(self):
        """Return dict of all form fields."""
        return {
            "request_date":       self.req_date.get().strip(),
            "org_name":           self.req_org.get().strip(),
            "training_requested": self.req_training.get().strip(),
            "requested_date":     self.req_req_date.get().strip(),
            "location":           self.req_location.get().strip(),
            "num_participants":   self.req_num_part.get().strip(),
            "participant_roles":  self.req_roles.get().strip(),
            "preferred_trainer":  self.req_trainer.get().strip(),
            "mileage_cost":       self.req_mileage.get().strip(),
            "other_travel_cost":  self.req_travel.get().strip(),
            "billing_contact":    self.req_bill_name.get().strip(),
            "billing_email":      self.req_bill_email.get().strip(),
            "billing_phone":      self.req_bill_phone.get().strip(),
            "additional_comments": self.req_comments.get("1.0", tk.END).strip(),
        }

    def _req_quote_changed(self, new_val=None):
        """Recalculate total whenever a cost field changes."""
        total = 0.0
        for _, _, c in self._req_line_items:
            try:
                total += float(c.get().replace("$", "").replace(",", "").strip() or 0)
            except ValueError:
                pass
        self.req_quote_total.config(text=f"${total:,.2f}")
        self._req_populate_quote_email()
        return True   # allow key press

    def _req_populate_quote_email(self):
        """Build a default quote email body from current form + line items."""
        if not hasattr(self, "req_quote_body"):
            return
        org    = self.req_org.get().strip() if hasattr(self, "req_org") else ""
        name   = self.req_bill_name.get().strip() if hasattr(self, "req_bill_name") else ""
        train  = self.req_training.get().strip() if hasattr(self, "req_training") else ""
        rdate  = self.req_req_date.get().strip() if hasattr(self, "req_req_date") else ""
        loc    = self.req_location.get().strip() if hasattr(self, "req_location") else ""

        lines_txt = ""
        total = 0.0
        if hasattr(self, "_req_line_items"):
            for de, dte, ce in self._req_line_items:
                d = de.get().strip(); dt = dte.get().strip(); c = ce.get().strip()
                if d or c:
                    cost_val = 0.0
                    try:
                        cost_val = float(c.replace("$","").replace(",","") or 0)
                    except ValueError:
                        pass
                    total += cost_val
                    lines_txt += f"\n  • {d}"
                    if dt:
                        lines_txt += f"  ({dt})"
                    if c:
                        lines_txt += f":  ${cost_val:,.2f}"

        greeting = f"Dear {name}," if name else "Dear Training Contact,"
        body = (
            f"{greeting}\n\n"
            f"Thank you for your interest in training from the "
            f"University of Kentucky Center on Trauma and Children.\n\n"
            f"Please find attached a training quote"
            + (f" for {org}" if org else "")
            + (f" for \"{train}\"" if train else "")
            + (f" on {rdate}" if rdate else "")
            + (f" at {loc}" if loc else "")
            + ".\n\n"
            + ("Quote Details:" + lines_txt + f"\n\nTotal:  ${total:,.2f}\n\n" if lines_txt else "")
            + "Please review the attached quote and let us know if you have any questions "
            "or would like to proceed. Quotes are valid for 90 days.\n\n"
            "Best regards,\n\n"
            "Joshua Fisherkeller, MSW\n"
            "Education and Training Manager\n"
            "University of Kentucky Center on Trauma and Children\n"
            "College of Medicine, Department of Psychiatry\n"
            "3470 Blazer Parkway Suite 100\n"
            "Lexington, Kentucky 40509\n"
            "(859) 218-6941\n"
            "http://www.uky.edu/CTAC/"
        )
        self.req_quote_body.delete("1.0", tk.END)
        self.req_quote_body.insert("1.0", body)

    def _req_send_to_sprang(self, test=False):
        """Generate the request PDF, email to Dr. Sprang (or test), save JSON + spreadsheet row."""
        data = self._req_get_form_data()
        if not data["org_name"]:
            messagebox.showerror("Missing Field", "Please enter the Organization name.")
            return
        if not data["training_requested"]:
            messagebox.showerror("Missing Field", "Please enter the Training Requested.")
            return

        try:
            # Ensure store dir exists
            os.makedirs(self._req_store, exist_ok=True)

            # Build unique ID
            safe_org = re.sub(r"[^\w]", "_", data["org_name"])[:30]
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            req_id = f"{ts}_{safe_org}"
            data["id"] = req_id
            data["status"] = "submitted"
            data["submitted_at"] = datetime.now().isoformat()
            data["line_items"] = []
            data["quote_total"] = ""
            data["quoted_at"] = None
            data["resolved_at"] = None
            data["outcome"] = None

            # Generate PDF
            pdf_path = os.path.join(self._req_store, f"request_{req_id}.pdf")
            generate_request_pdf(data, pdf_path)

            # For real sends: save JSON and update Master Training List
            if not test:
                _req_save_json(self._req_store, data)
                mtl = self.mtl_path.get().strip() if hasattr(self, "mtl_path") else ""
                if mtl and os.path.isfile(mtl):
                    _req_append_to_master(mtl, data)

            # Determine recipient + CC
            org   = data["org_name"]
            rdate = data["requested_date"]
            subj  = f"Training Request – {org} – {rdate or data['request_date']}"
            if test:
                to_addr = TEST_EMAIL
                cc_addr = None
            else:
                to_addr = "sprang@uky.edu"
                cc_addr = TEST_EMAIL   # always CC Josh on real sends

            # Build email body
            train   = data["training_requested"]
            loc     = data["location"]
            trainer = data["preferred_trainer"]
            mileage = data["mileage_cost"]
            travel  = data["other_travel_cost"]
            parts   = data["num_participants"]
            roles   = data["participant_roles"]
            bill    = data["billing_contact"]
            notes   = data["additional_comments"]

            body_lines = [
                "<b>New Training Request</b>",
                "",
                f"<b>Organization:</b> {org}",
                f"<b>Training Requested:</b> {train}",
                f"<b>Requested Date:</b> {rdate}",
                f"<b>Location:</b> {loc}",
                "",
                f"<b>Participants:</b> {parts}",
                f"<b>Roles:</b> {roles}",
                "",
                f"<b>Preferred Trainer:</b> {trainer}",
                f"<b>Mileage Cost:</b> {mileage}",
                f"<b>Other Travel:</b> {travel}",
                "",
                f"<b>Billing Contact:</b> {bill}",
            ]
            if notes:
                body_lines += ["", f"<b>Notes:</b> {notes}"]
            if test:
                body_lines = [
                    "<i>[TEST EMAIL — would be sent to sprang@uky.edu with CC to jafish0@uky.edu]</i>",
                    "",
                ] + body_lines
            body_lines += [
                "",
                "Please see the attached PDF for the full request form.",
                "",
                "—",
                "Joshua Fisherkeller, MSW",
                "Education and Training Manager",
                "University of Kentucky Center on Trauma and Children",
            ]
            html_body = "<br>".join(body_lines)

            # Send via Outlook — use CC for real sends
            send_outlook_email(to_addr, subj, html_body, pdf_path,
                               html=True, cc=cc_addr)

            if not test:
                self._req_current_id = req_id
                self._req_refresh_pending_list()
                self.req_status.config(
                    text=f"✔ Sent to Dr. Sprang (CC: {TEST_EMAIL}). ID: {req_id[:16]}…")
                messagebox.showinfo("Request Sent",
                    f'Training request for "{org}" sent to Dr. Sprang.\n'
                    f"You were CC'd at {TEST_EMAIL}.\n\n"
                    f"PDF saved to:\n{pdf_path}")
            else:
                self.req_status.config(
                    text=f"✔ Test sent to {TEST_EMAIL}.")
                messagebox.showinfo("Test Sent",
                    f'Test request PDF sent to {TEST_EMAIL}.\n\n'
                    f"PDF saved to:\n{pdf_path}\n\n"
                    "(No JSON saved, no spreadsheet updated.)")

        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def _req_preview_send_quote(self):
        """Open a review window so Josh can edit and send the quote email."""
        to_addr = self.req_quote_to.get().strip()
        if not to_addr:
            messagebox.showerror("Missing", "Please enter a Send To email address.")
            return

        # Gather line items for quote PDF
        line_items = []
        for de, dte, ce in self._req_line_items:
            d = de.get().strip(); dt = dte.get().strip(); c = ce.get().strip()
            if d or c:
                line_items.append({"desc": d, "date": dt, "cost": c})
        if not line_items:
            messagebox.showerror("Missing", "Please enter at least one quote line item.")
            return

        data = self._req_get_form_data()
        data["line_items"] = line_items
        total = sum(
            float(li["cost"].replace("$","").replace(",","") or 0)
            for li in line_items
            if li.get("cost")
        )
        data["quote_total"] = f"{total:,.2f}"

        # Generate quote PDF
        try:
            os.makedirs(self._req_store, exist_ok=True)
            safe_org = re.sub(r"[^\w]", "_", data["org_name"])[:30]
            pdf_name = f"quote_{safe_org}.pdf"
            pdf_path = os.path.join(self._req_store, pdf_name)
            generate_quote_pdf(data, pdf_path)
        except Exception as exc:
            messagebox.showerror("PDF Error", str(exc))
            return

        # Build preview window
        body_text = self.req_quote_body.get("1.0", tk.END).strip()
        win = tk.Toplevel(self)
        win.title("Review & Send Quote")
        win.geometry("740x580")
        win.grab_set()

        ttk.Label(win, text="Review the email below, edit if needed, then click Send.",
                  foreground="navy").pack(padx=14, pady=(10, 4), anchor="w")

        frm_hdr = ttk.Frame(win)
        frm_hdr.pack(fill="x", padx=14, pady=2)
        ttk.Label(frm_hdr, text="To:").grid(row=0, column=0, sticky="w", padx=(0,6))
        to_ent = ttk.Entry(frm_hdr, width=50)
        to_ent.insert(0, to_addr)
        to_ent.grid(row=0, column=1, sticky="w")
        ttk.Label(frm_hdr, text="Attachment:").grid(row=1, column=0, sticky="w", padx=(0,6), pady=2)
        ttk.Label(frm_hdr, text=pdf_name, foreground="blue").grid(row=1, column=1, sticky="w")

        ttk.Label(win, text="Email Body:").pack(padx=14, anchor="w")
        body_frame = ttk.Frame(win)
        body_frame.pack(fill="both", expand=True, padx=14, pady=4)
        body_sb = ttk.Scrollbar(body_frame)
        body_sb.pack(side="right", fill="y")
        body_edit = tk.Text(body_frame, wrap="word", yscrollcommand=body_sb.set)
        body_edit.pack(fill="both", expand=True)
        body_sb.config(command=body_edit.yview)
        body_edit.insert("1.0", body_text)

        btn_row = ttk.Frame(win)
        btn_row.pack(fill="x", padx=14, pady=8)

        def _do_send():
            final_to = to_ent.get().strip()
            final_body = body_edit.get("1.0", tk.END).strip()
            try:
                send_outlook_email(
                    final_to,
                    f"Training Quote – {data['org_name']}",
                    md_to_html(final_body),
                    pdf_path, html=True
                )
                # Update JSON with quote info
                if self._req_current_id:
                    existing = _req_load_json(self._req_store, self._req_current_id)
                    if existing:
                        existing["status"] = "quoted"
                        existing["line_items"] = line_items
                        existing["quote_total"] = data["quote_total"]
                        existing["quoted_at"] = datetime.now().isoformat()
                        _req_save_json(self._req_store, existing)
                        # Update Master Training List
                        mtl = self.mtl_path.get().strip() if hasattr(self, "mtl_path") else ""
                        if mtl and os.path.isfile(mtl):
                            _req_update_master_quote(
                                mtl, existing["org_name"],
                                data["quote_total"],
                                datetime.now().strftime("%m/%d/%Y")
                            )
                self._req_refresh_pending_list()
                self.req_quote_status.config(text=f"✔ Quote sent to {final_to}")
                win.destroy()
                messagebox.showinfo("Quote Sent",
                    f"Quote emailed to {final_to}.\nPDF: {pdf_path}")
            except Exception as exc:
                messagebox.showerror("Send Error", str(exc), parent=win)

        ttk.Button(btn_row, text="✉  Send Quote", command=_do_send).pack(
            side="left", padx=4)
        ttk.Button(btn_row, text="Cancel", command=win.destroy).pack(
            side="left", padx=4)

    def _req_mark_resolved(self, outcome):
        """Mark the loaded request as accepted/declined/cancelled."""
        if not self._req_current_id:
            messagebox.showerror("No Request Loaded",
                "Load a pending request first using 'Load Pending Request'.")
            return
        label = {"accepted": "Accepted", "declined": "Declined",
                 "cancelled": "Cancelled"}.get(outcome, outcome)
        if not messagebox.askyesno("Confirm",
                f"Mark this request as '{label}'?"):
            return
        existing = _req_load_json(self._req_store, self._req_current_id)
        if not existing:
            messagebox.showerror("Error", "Could not find the JSON record for this request.")
            return
        existing["status"] = outcome
        existing["resolved_at"] = datetime.now().isoformat()
        existing["outcome"] = outcome
        _req_save_json(self._req_store, existing)

        # Move in Master Training List
        mtl = self.mtl_path.get().strip() if hasattr(self, "mtl_path") else ""
        if mtl and os.path.isfile(mtl):
            _req_resolve_in_master(mtl, existing, outcome)

        self._req_current_id = None
        self._req_refresh_pending_list()
        self.req_quote_status.config(text=f"✔ Marked as {label}.")
        messagebox.showinfo("Updated", f"Request marked as {label} and moved to Resolved.")

    def _req_start_over(self):
        """Clear all request fields."""
        for w in (self.req_org, self.req_training, self.req_req_date,
                  self.req_location, self.req_num_part, self.req_roles,
                  self.req_mileage, self.req_travel, self.req_bill_name,
                  self.req_bill_email, self.req_bill_phone, self.req_quote_to):
            w.delete(0, tk.END)
        self.req_date.delete(0, tk.END)
        self.req_date.insert(0, datetime.now().strftime("%m/%d/%Y"))
        self.req_trainer.set("")
        self.req_comments.delete("1.0", tk.END)
        for de, dte, ce in self._req_line_items:
            de.delete(0, tk.END); dte.delete(0, tk.END); ce.delete(0, tk.END)
        self.req_quote_total.config(text="$0.00")
        self._req_current_id = None
        self._req_mode.set(1)
        self._req_mode_toggle()
        self.req_status.config(text="")
        self.req_quote_status.config(text="")
        self._req_populate_quote_email()


# ---------- Learning Collaborative helpers ----------

def _find_importid_row_idx(raw: pd.DataFrame):
    for idx in range(min(6, len(raw))):
        row = raw.iloc[idx]
        if row.apply(lambda x: isinstance(x, str) and "importid" in x.lower()).any():
            return idx
    return None


def _qualtrics_name_email_idx(labels: list, importids: list):
    """Return (name_idx, email_idx) from Qualtrics label/importid rows."""
    name_idx = None; best_n = (9999, 9999, 9999)
    for i, lab in enumerate(labels):
        s = str(lab).strip(); sl = s.lower()
        if "name" in sl and "email" not in sl:
            recip = 1 if "recipient" in sl else 0
            tup = (recip, 0 if sl == "name" else 1, len(sl))
            if tup < best_n: best_n, name_idx = tup, i
    email_idx = None; best_e = (9999, 9999, 9999, 9999)
    for i, lab in enumerate(labels):
        s = str(lab).strip(); sl = s.lower()
        if "email" in sl:
            recip = 1 if "recipient" in sl else 0
            confirm = 1 if "confirm" in sl else 0
            techname = str(importids[i]).strip().lower() if i < len(importids) else ""
            is_q = 0 if re.match(r"q\d+", techname) else 1
            tup = (recip, confirm, is_q, len(sl))
            if tup < best_e: best_e, email_idx = tup, i
    return name_idx, email_idx


def _lc_count_psych_for_date(si_csv_path: str, date_str: str) -> int:
    """Count psychology-role participants in a single LC session date.

    Loads the raw sign-in CSV, filters to rows whose StartDate/RecordedDate
    starts with date_str (YYYY-MM-DD), then delegates to count_psychologists.
    Returns 0 if the role column is absent or the file can't be read.
    """
    try:
        raw = pd.read_csv(si_csv_path, dtype=str, keep_default_na=False)
        # Qualtrics CSVs have 2-3 header rows; drop until we hit the data.
        # Normalize the date column so both ISO ('2026-01-20 …') and
        # US-style ('1/20/2026 …') exports work.
        date_col = raw.columns[0]
        norm = raw[date_col].apply(_normalize_qualtrics_date)
        data = raw[norm == date_str].copy()
        return count_psychologists(data)
    except Exception:
        return 0


def _normalize_qualtrics_date(s) -> str:
    """Normalize a Qualtrics start-date cell to 'YYYY-MM-DD'.

    Accepts Qualtrics' default ISO format ('2026-01-20 09:58:00') as well
    as the US-style formats Excel or a non-default Qualtrics download
    option can produce ('1/20/2026 9:58', '01/20/2026', '1/20/26 9:58',
    '2026/01/20'). Returns '' when the cell can't be parsed."""
    if not isinstance(s, str):
        return ""
    s = s.strip()
    if not s:
        return ""
    # Already ISO YYYY-MM-DD (with or without a time suffix)
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # Otherwise: take the portion before whitespace (drop any time)
    date_part = s.split()[0]
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(date_part, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def parse_lc_signin_sessions(path: str) -> pd.DataFrame:
    """Parse an LC sign-in Qualtrics CSV.
    Returns DataFrame with columns: DATE (YYYY-MM-DD), NAME, EMAIL.
    One row per submission — same person may appear across multiple dates."""
    raw = pd.read_csv(path, dtype=str, header=None, keep_default_na=False)
    importid_idx = _find_importid_row_idx(raw)
    if importid_idx is None:
        raise ValueError("Could not detect Qualtrics format in sign-in CSV.")
    labels = list(raw.iloc[importid_idx - 1])
    importids = list(raw.iloc[importid_idx])
    data_start = importid_idx + 1
    name_idx, email_idx = _qualtrics_name_email_idx(labels, importids)
    if email_idx is None:
        raise ValueError("Could not find email column in sign-in CSV.")
    data = raw.iloc[data_start:].reset_index(drop=True)
    df = pd.DataFrame({
        "DATE":  data.iloc[:, 0].apply(_normalize_qualtrics_date),
        "NAME":  data.iloc[:, name_idx].fillna("") if name_idx is not None else "",
        "EMAIL": data.iloc[:, email_idx].fillna(""),
    })
    df = _smart_strip_df(df)
    df["NAME"]  = df["NAME"].apply(smart_title)
    df["EMAIL"] = df["EMAIL"].str.lower()
    df = df[df["DATE"].str.match(r"\d{4}-\d{2}-\d{2}", na=False)]
    df = df[df["EMAIL"].apply(looks_like_email)]
    df = df.reset_index(drop=True)
    if df.empty:
        raise ValueError(
            "No valid dated rows found in the Sign-In CSV.\n"
            "Check that the first column contains dates like "
            "'2026-01-20 09:58:00' or '1/20/2026 9:58', and that the "
            "email column is populated.")
    return df


def parse_lc_signout_sessions(path: str) -> pd.DataFrame:
    """Parse an LC sign-out Qualtrics CSV.
    Returns DataFrame with columns: DATE, NAME, EMAIL, WANTS_CERT (bool)."""
    raw = pd.read_csv(path, dtype=str, header=None, keep_default_na=False)
    importid_idx = _find_importid_row_idx(raw)
    if importid_idx is None:
        raise ValueError("Could not detect Qualtrics format in sign-out CSV.")
    labels   = list(raw.iloc[importid_idx - 1])
    importids = list(raw.iloc[importid_idx])
    data_start = importid_idx + 1
    name_idx, email_idx = _qualtrics_name_email_idx(labels, importids)
    # Find WANTS_CERT column — QID17 ImportId or label "CE or a certificate"
    cert_idx = None
    for i, imp in enumerate(importids):
        if isinstance(imp, str) and "QID17" in imp.upper():
            cert_idx = i; break
    if cert_idx is None:
        for i, lab in enumerate(labels):
            sl = str(lab).strip().lower()
            if "certificate" in sl and ("would you like" in sl or "receive ce" in sl):
                cert_idx = i; break
    data = raw.iloc[data_start:].reset_index(drop=True)
    if cert_idx is not None:
        cert_series = data.iloc[:, cert_idx].fillna("").str.strip().str.lower().isin(["yes"])
    else:
        cert_series = pd.Series([True] * len(data))
    df = pd.DataFrame({
        "DATE":       data.iloc[:, 0].apply(_normalize_qualtrics_date),
        "NAME":       data.iloc[:, name_idx].fillna("") if name_idx is not None else "",
        "EMAIL":      data.iloc[:, email_idx].fillna("") if email_idx is not None else "",
        "WANTS_CERT": cert_series,
    })
    df = _smart_strip_df(df)
    df["NAME"]  = df["NAME"].apply(smart_title)
    df["EMAIL"] = df["EMAIL"].str.lower()
    df = df[df["DATE"].str.match(r"\d{4}-\d{2}-\d{2}", na=False)]
    df = df.reset_index(drop=True)
    if df.empty:
        raise ValueError(
            "No valid dated rows found in the Sign-Out CSV.\n"
            "Check that the first column contains dates like "
            "'2026-01-20 09:58:00' or '1/20/2026 9:58'.")
    return df


def get_lc_session_dates(si_df: pd.DataFrame, so_df) -> list:
    """Return sorted list of (date_str, likely_real, sign_in_count) tuples.

    When both CSVs are provided, likely_real = True when the date appears in
    both and has >= 3 sign-in entries.
    When so_df is None (sign-in-only mode), likely_real = True when the date
    has >= 3 sign-in entries."""
    si_counts = si_df.groupby("DATE").size()
    so_dates  = set(so_df["DATE"].unique()) if so_df is not None else None
    results = []
    for date, count in si_counts.items():
        if so_dates is None:
            in_both = True   # sign-in-only: no sign-out cross-check
        else:
            in_both = date in so_dates
        likely_real = in_both and count >= 3
        results.append((date, likely_real, int(count)))
    return sorted(results, key=lambda x: x[0])


def build_lc_attendance(si_df: pd.DataFrame, so_df,
                         sessions: list, excluded: list) -> list:
    """Build per-participant attendance data with fuzzy name matching fallback.

    sessions : list of (date_str YYYY-MM-DD, hours float)
    excluded : list of name strings (case-insensitive)
    so_df    : sign-out DataFrame, or None for sign-in-only mode (in which
               case attendance is counted from sign-ins alone)

    Returns list of dicts:
      {name, email, hours_attended, hours_total,
       session_data: [(date, hours, signed_in, signed_out, counted), ...],
       fuzzy_matched: bool,
       signin_only: bool}
    """
    signin_only = so_df is None
    excluded_lower = {e.strip().lower() for e in excluded if e.strip()}
    hours_total    = sum(h for _, h in sessions)

    # Build sign-in lookup: date -> set of emails, and name/email maps
    si_by_date: dict = {}
    si_name_by_email: dict = {}
    si_email_by_name_lower: dict = {}   # for fuzzy fallback
    for _, row in si_df.iterrows():
        nm = row["NAME"]; em = row["EMAIL"]; dt = row["DATE"]
        if nm.lower() in excluded_lower or not looks_like_email(em): continue
        si_by_date.setdefault(dt, set()).add(em)
        si_name_by_email[em] = nm
        si_email_by_name_lower[nm.strip().lower()] = em

    # Fuzzy canonical email: map sign-out email → canonical sign-in email
    def _canonical(so_em, so_nm):
        if so_em in si_name_by_email:
            return so_em, False          # exact email match
        si_names = list(si_email_by_name_lower.keys())
        if not si_names:
            return so_em, False
        matches = difflib.get_close_matches(so_nm.strip().lower(), si_names, n=1, cutoff=0.82)
        if matches:
            return si_email_by_name_lower[matches[0]], True   # fuzzy match
        return so_em, False

    # Build sign-out lookup: date -> set of canonical emails that want cert
    so_by_date: dict = {}
    fuzzy_matched_emails: set = set()
    if not signin_only:
        for _, row in so_df.iterrows():
            nm = row["NAME"]; em = row["EMAIL"]; dt = row["DATE"]
            wants = bool(row.get("WANTS_CERT", True))
            if nm.lower() in excluded_lower or not looks_like_email(em): continue
            canon_em, was_fuzzy = _canonical(em, nm)
            if was_fuzzy:
                fuzzy_matched_emails.add(canon_em)
            if wants:
                so_by_date.setdefault(dt, set()).add(canon_em)

    # All unique participant emails (from sign-in)
    all_emails = set()
    for s in si_by_date.values():
        all_emails |= s

    participants = []
    for em in sorted(all_emails):
        name = si_name_by_email.get(em, em)
        session_data = []
        hours_attended = 0.0
        for (date, hours) in sessions:
            signed_in  = em in si_by_date.get(date, set())
            if signin_only:
                # Treat sign-in as attendance; no sign-out CSV to cross-check
                signed_out = signed_in
                counted    = signed_in
            else:
                signed_out = em in so_by_date.get(date, set())
                counted    = signed_in and signed_out
            if counted:
                hours_attended += hours
            session_data.append((date, hours, signed_in, signed_out, counted))
        participants.append({
            "name":           name,
            "email":          em,
            "hours_attended": hours_attended,
            "hours_total":    hours_total,
            "session_data":   session_data,
            "fuzzy_matched":  em in fuzzy_matched_emails,
            "signin_only":    signin_only,
        })
    return participants


def _lc_date_range(sessions: list) -> str:
    """Return a human-readable date range string, e.g. '11/12/2025 – 3/17/2026'."""
    if not sessions: return ""
    dts = []
    for date_str, _ in sessions:
        try: dts.append(datetime.strptime(date_str, "%Y-%m-%d"))
        except Exception: pass
    if not dts: return ""
    fmt = "%#m/%#d/%Y" if os.name == "nt" else "%-m/%-d/%Y"
    if len(dts) == 1: return dts[0].strftime(fmt)
    return f"{min(dts).strftime(fmt)} \u2013 {max(dts).strftime(fmt)}"


def build_attendance_html(participant: dict, dates_fmt: list) -> str:
    """Return an HTML table showing per-session attendance for one participant.

    In sign-in-only mode (participant['signin_only'] is True) the table omits
    the Signed Out column and presents a single Attended column instead.

    Roster mode (precomputed app roster) has no per-session data — render a
    one-line summary instead of an empty table."""
    if not participant.get("session_data"):
        fmt_n = lambda n: f"{float(n):g}"
        return (
            "<p style='font-family:Arial,sans-serif;font-size:13px;margin:8px 0;'>"
            f"You attended <strong>{fmt_n(participant['hours_attended'])}</strong> of "
            f"<strong>{fmt_n(participant['hours_total'])}</strong> possible hours."
            "</p>"
        )
    signin_only = bool(participant.get("signin_only"))
    rows_html = ""
    for i, (date, hours, si, so, counted) in enumerate(participant["session_data"]):
        hrs_str  = (f"{hours:g} hr{'s' if hours != 1 else ''}"
                    if counted else "&mdash;")
        date_fmt = dates_fmt[i] if i < len(dates_fmt) else date
        if signin_only:
            attended_mark = "&#10003;" if si else "&mdash;"
            rows_html += (
                f"<tr>"
                f"<td style='padding:3px 10px;border:1px solid #ccc;'>{date_fmt}</td>"
                f"<td style='padding:3px 10px;border:1px solid #ccc;text-align:center;'>{attended_mark}</td>"
                f"<td style='padding:3px 10px;border:1px solid #ccc;text-align:center;'>{hrs_str}</td>"
                f"</tr>\n"
            )
        else:
            si_mark = "&#10003;" if si else "&mdash;"
            so_mark = "&#10003;" if so else "&mdash;"
            rows_html += (
                f"<tr>"
                f"<td style='padding:3px 10px;border:1px solid #ccc;'>{date_fmt}</td>"
                f"<td style='padding:3px 10px;border:1px solid #ccc;text-align:center;'>{si_mark}</td>"
                f"<td style='padding:3px 10px;border:1px solid #ccc;text-align:center;'>{so_mark}</td>"
                f"<td style='padding:3px 10px;border:1px solid #ccc;text-align:center;'>{hrs_str}</td>"
                f"</tr>\n"
            )
    h_att = participant["hours_attended"]
    h_tot = participant["hours_total"]
    fmt_n = lambda n: f"{float(n):g}"
    if signin_only:
        rows_html += (
            f"<tr style='font-weight:bold;background:#f7f7f7;'>"
            f"<td style='padding:3px 10px;border:1px solid #ccc;'>Total</td>"
            f"<td style='padding:3px 10px;border:1px solid #ccc;'></td>"
            f"<td style='padding:3px 10px;border:1px solid #ccc;text-align:center;'>"
            f"{fmt_n(h_att)} / {fmt_n(h_tot)} hrs</td>"
            f"</tr>\n"
        )
        header = (
            "<tr style='background:#e8e8e8;'>"
            "<th style='padding:4px 10px;border:1px solid #ccc;text-align:left;'>Session Date</th>"
            "<th style='padding:4px 10px;border:1px solid #ccc;'>Attended</th>"
            "<th style='padding:4px 10px;border:1px solid #ccc;'>Hours Credited</th>"
            "</tr>\n"
        )
    else:
        rows_html += (
            f"<tr style='font-weight:bold;background:#f7f7f7;'>"
            f"<td style='padding:3px 10px;border:1px solid #ccc;'>Total</td>"
            f"<td style='padding:3px 10px;border:1px solid #ccc;'></td>"
            f"<td style='padding:3px 10px;border:1px solid #ccc;'></td>"
            f"<td style='padding:3px 10px;border:1px solid #ccc;text-align:center;'>"
            f"{fmt_n(h_att)} / {fmt_n(h_tot)} hrs</td>"
            f"</tr>\n"
        )
        header = (
            "<tr style='background:#e8e8e8;'>"
            "<th style='padding:4px 10px;border:1px solid #ccc;text-align:left;'>Session Date</th>"
            "<th style='padding:4px 10px;border:1px solid #ccc;'>Signed In</th>"
            "<th style='padding:4px 10px;border:1px solid #ccc;'>Signed Out</th>"
            "<th style='padding:4px 10px;border:1px solid #ccc;'>Hours Credited</th>"
            "</tr>\n"
        )
    return (
        "<table style='border-collapse:collapse;font-family:Arial,sans-serif;"
        "font-size:13px;margin:8px 0;'>\n"
        + header + rows_html + "</table>"
    )


def generate_lc_cert(name: str, hours_attended, hours_total,
                      event_name: str, trainer: str, date_range: str,
                      template: str, outdir: str,
                      location: str = "", eila: str = "",
                      ceu_vars: dict = None) -> str:
    """Generate one LC certificate. Returns path to the saved file (PDF if possible)."""
    fmt_n = lambda n: f"{float(n):g}"
    mapping = {
        "<<NAME>>":            name,
        "<<TRAINING>>":        event_name,
        "<<HOURS_ATTENDED>>":  fmt_n(hours_attended),
        "<<HOURS_TOTAL>>":     fmt_n(hours_total),
        "<<TRAINER>>":         trainer,
        "<<DATE>>":            date_range,
        "<<DATES>>":           date_range,
        "<<HOURS>>":           fmt_n(hours_attended),   # fallback for standard template
        "<<LOCATION>>":        location,
        "<<EILA>>":            eila,
    }
    if ceu_vars:
        mapping.update(ceu_vars)
    os.makedirs(outdir, exist_ok=True)
    doc = Document(template)
    replace_placeholders_everywhere(doc, mapping)
    safe_name     = re.sub(r"[^A-Za-z0-9._ -]", "", name)
    safe_training = safe_filename(event_name)
    base = f"{safe_name} - {safe_training}".strip().replace("  ", " ")
    out_docx = os.path.join(outdir, base + ".docx")
    doc.save(out_docx)
    out_file = out_docx
    if docx2pdf_convert is not None:
        try:
            out_pdf = os.path.join(outdir, base + ".pdf")
            docx2pdf_convert(out_docx, out_pdf)
            if os.path.isfile(out_pdf): out_file = out_pdf
        except Exception:
            pass
    return out_file


# ---------- Attendance Ingest Layer (canonical form) ----------
#
# Phase 1 of the flexible-ingest refactor.
#
# This layer sits ALONGSIDE the existing parsers (load_csv,
# parse_lc_signin_sessions, parse_lc_signout_sessions) and produces a single
# canonical shape for any supported attendance source. No existing caller is
# wired through this layer yet — Phase 1 is pure infrastructure. The
# Qualtrics path used by the Standard and LC tabs is unchanged.
#
# Adding a new format (Phase 2+) means writing a Profile subclass and
# appending it to DEFAULT_PROFILES.

CANONICAL_COLUMNS = (
    "NAME",          # str, required
    "EMAIL",         # str, required (may be empty)
    "ROLE",          # str, optional
    "AGENCY",        # str, optional
    "ATTENDED",      # bool, optional (None = unknown)
    "HOURS",         # float, optional (None = filled by UI later)
    "SESSION_DATE",  # str YYYY-MM-DD, optional
    "SESSION_LABEL", # str, optional ("Session 1", "Morning", etc.)
    "WANTS_CERT",    # bool, optional (sign-out CSVs)
)


@dataclass
class CanonicalAttendance:
    """A normalized representation of an attendance source.

    All ingest profiles parse into this shape so downstream certificate /
    email code can consume any source uniformly.

    Attributes
    ----------
    rows : pd.DataFrame
        One row per attendance record. Must include NAME and EMAIL; other
        columns from CANONICAL_COLUMNS are optional. Profile-specific extra
        columns are allowed and ignored by downstream code.
    granularity : str
        "per_signin"     — one row per sign-in event (Qualtrics shape).
        "per_participant" — one row per person with per-session columns
                            (roster shape; introduced in Phase 2).
    training_title : Optional[str]
        Inferred from filename or explicit metadata; may be None.
    session_columns : list
        For per_participant sources: list of (label, date|None,
        default_hours|None) describing per-session columns in the source.
    source_path : str
        Original file path on disk.
    profile_used : str
        Name of the Profile that produced this object.
    confidence : float
        0.0-1.0 confidence the chosen profile fits the file.
    warnings : list[str]
        Human-readable issues to surface in the mapping confirmation UI
        (missing emails, duplicates, fuzzy domain matches, etc.).
    """
    rows: pd.DataFrame
    granularity: str = "per_signin"
    training_title: Optional[str] = None
    session_columns: list = field(default_factory=list)
    source_path: str = ""
    profile_used: str = ""
    confidence: float = 0.0
    warnings: list = field(default_factory=list)

    def is_empty(self) -> bool:
        return self.rows is None or self.rows.empty

    def __len__(self) -> int:
        return 0 if self.rows is None else len(self.rows)


class Profile:
    """Base class for an attendance source profile.

    Subclasses must implement:
      name          : human-readable identifier (kebab/snake case)
      detect(path)  : float 0.0-1.0 confidence this profile fits the file
      parse(path)   : return a CanonicalAttendance
    """
    name: str = "base"
    granularity: str = "per_signin"

    def detect(self, path: str) -> float:
        return 0.0

    def parse(self, path: str) -> CanonicalAttendance:
        raise NotImplementedError


def _title_from_filename(path: str) -> str:
    """Best-effort extraction of a training title from a Qualtrics-style
    filename. Strips trailing Qualtrics date stamps, "+" URL-encoded spaces,
    and Windows-copy suffixes like "(1)"."""
    stem = os.path.splitext(os.path.basename(path))[0]
    stem = stem.replace("+", " ")
    # Trim trailing "(1)" Windows-copy suffix FIRST so the date trim below
    # can anchor to the real end of the title.
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
    # Trim trailing "_Month Day, YYYY_HH.MM" stamp common in Qualtrics exports.
    # Use [A-Za-z]+ (not \w+) so the month name doesn't greedily eat
    # underscore-joined words like "Sessions_May".
    stem = re.sub(r"[_ ]+[A-Za-z]+\s+\d{1,2},\s*\d{4}[_ ]?\d{0,2}\.?\d*$", "", stem)
    return stem.strip()


def _filename_hints(path: str):
    """Lowercased filename plus a set of normalized tokens for matching."""
    lower = os.path.basename(path).lower()
    norm  = re.sub(r"[+_\-.]+", " ", lower)
    return lower, norm


class QualtricsStandardProfile(Profile):
    """Qualtrics export consumed by the Standard tab.

    Wraps the existing load_csv() function so behavior is unchanged. The
    rows DataFrame keeps the same NAME/EMAIL/ROLE shape load_csv produces
    today; no per-row session date (the trainer enters one date in the UI)."""
    name = "qualtrics_standard"
    granularity = "per_signin"

    def detect(self, path):
        if not path.lower().endswith(".csv"):
            return 0.0
        if _detect_is_qualtrics_csv(path):
            # Lower than the LC-specific profiles so filename hints win when
            # they fit; this is the catch-all for Qualtrics CSVs.
            return 0.5
        return 0.0

    def parse(self, path):
        df = load_csv(path)
        return CanonicalAttendance(
            rows=df,
            granularity="per_signin",
            training_title=_title_from_filename(path),
            source_path=path,
            profile_used=self.name,
            confidence=self.detect(path),
        )


class QualtricsLcSigninProfile(Profile):
    """Qualtrics LC sign-in export. One row per sign-in, per-row SESSION_DATE.
    Wraps parse_lc_signin_sessions()."""
    name = "qualtrics_lc_signin"
    granularity = "per_signin"

    def detect(self, path):
        if not path.lower().endswith(".csv") or not _detect_is_qualtrics_csv(path):
            return 0.0
        _, norm = _filename_hints(path)
        # "sign in" wins; reject if filename also says "sign out"
        if "sign in" in norm or "signin" in norm:
            if "sign out" in norm or "signout" in norm:
                return 0.0
            return 0.95
        return 0.0

    def parse(self, path):
        df = parse_lc_signin_sessions(path)
        rows = df.rename(columns={"DATE": "SESSION_DATE"})
        return CanonicalAttendance(
            rows=rows,
            granularity="per_signin",
            training_title=_title_from_filename(path),
            source_path=path,
            profile_used=self.name,
            confidence=self.detect(path),
        )


class QualtricsLcSignoutProfile(Profile):
    """Qualtrics LC sign-out export. Wraps parse_lc_signout_sessions()."""
    name = "qualtrics_lc_signout"
    granularity = "per_signin"

    def detect(self, path):
        if not path.lower().endswith(".csv") or not _detect_is_qualtrics_csv(path):
            return 0.0
        _, norm = _filename_hints(path)
        if "sign out" in norm or "signout" in norm:
            return 0.95
        return 0.0

    def parse(self, path):
        df = parse_lc_signout_sessions(path)
        rows = df.rename(columns={"DATE": "SESSION_DATE"})
        return CanonicalAttendance(
            rows=rows,
            granularity="per_signin",
            training_title=_title_from_filename(path),
            source_path=path,
            profile_used=self.name,
            confidence=self.detect(path),
        )


# Per-participant roster constants

# Attendance values that count as "this person attended this session".
# Compared after lower() + strip().
_YES_LIKE = frozenset({"yes", "y", "x", "✓", "✔", "true", "1", "present", "attended"})

# Header pattern for a per-session attendance column.
# Tolerates typos like "Sessoion 2" via the 'sess' prefix; also accepts
# "Day N" and "Module N" patterns.
_SESSION_COL_RE = re.compile(
    r"^\s*(?:sess\w*|day|module)\s*\d+\s*$", re.IGNORECASE
)


def _is_yes_like(v) -> bool:
    """Return True if `v` represents attendance (yes / y / x / ✓ / true /
    1 / present / attended, plus Python True and non-zero numerics)."""
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if isinstance(v, float) and pd.isna(v):
            return False
        return v != 0
    s = str(v).strip().lower()
    return s in _YES_LIKE


def _looks_like_unnamed(col) -> bool:
    """Detect pandas-generated placeholder column names like 'Unnamed: 7'.
    These appear when the source file has empty header cells; matching on
    them was the false-positive that caused the MTSS '0-sent' bug."""
    return isinstance(col, str) and col.strip().lower().startswith("unnamed")


def _find_roster_columns(df: pd.DataFrame):
    """Locate (name_col, email_col, session_cols) in a roster DataFrame.

    Raises ValueError if Name or Email cannot be identified. session_cols
    may be empty — the caller decides whether that's fatal."""
    name_col = email_col = None
    for c in df.columns:
        if _looks_like_unnamed(c):
            continue
        lc = str(c).strip().lower()
        if name_col is None and "name" in lc and "username" not in lc:
            name_col = c
        if email_col is None and "email" in lc and "confirm" not in lc:
            email_col = c
    if name_col is None:
        raise ValueError(
            "Could not find a 'Name' column in the roster. "
            "The header row needs a column whose label includes 'name'.")
    if email_col is None:
        raise ValueError(
            "Could not find an 'Email' column in the roster. "
            "The header row needs a column whose label includes 'email'.")
    session_cols = [c for c in df.columns
                    if not _looks_like_unnamed(c) and isinstance(c, str)
                    and _SESSION_COL_RE.match(c)]
    return name_col, email_col, session_cols


class PerParticipantRosterProfile(Profile):
    """Per-participant attendance roster (.xlsx, one row per person).

    Detection: an .xlsx whose first sheet's header row contains Name and
    Email plus one or more per-session columns ("Session N", "Day N",
    "Module N"; typos like "Sessoion 2" are tolerated).

    Parsing: returns long-form canonical rows — one row per (participant ×
    attended session) — with SESSION_LABEL set to the source column
    header. Hours and "# sessions attended" columns, if present, are
    intentionally ignored; hours come from the UI on certificate issue.

    Yes-like attendance values: yes / y / x / ✓ / true / 1 / present /
    attended (case-insensitive, trimmed). Anything else means absent."""
    name = "per_participant_roster"
    granularity = "per_participant"

    def detect(self, path):
        if not path.lower().endswith(".xlsx"):
            return 0.0
        try:
            df = pd.read_excel(path, sheet_name=0, dtype=object, nrows=30)
        except Exception:
            return 0.0
        try:
            _, _, session_cols = _find_roster_columns(df)
        except ValueError:
            return 0.0
        if not session_cols:
            return 0.0
        # Confirm at least one session column has mostly yes-like values.
        yes_ratio = 0.0
        for c in session_cols:
            col = df[c].dropna()
            if len(col) == 0:
                continue
            r = sum(1 for v in col if _is_yes_like(v)) / max(1, len(col))
            if r > yes_ratio:
                yes_ratio = r
        if yes_ratio < 0.2:
            return 0.4  # session-shaped columns but few yes-like values
        return 0.9

    def parse(self, path):
        df = pd.read_excel(path, sheet_name=0, dtype=object)
        df = _smart_strip_df(df)
        name_col, email_col, session_cols = _find_roster_columns(df)
        if not session_cols:
            raise ValueError(
                "No per-session columns found. Expected headers like "
                "'Session 1', 'Session 2', 'Day 1', or 'Module 1'.")

        # Drop fully-blank rows
        df = df.dropna(how="all").reset_index(drop=True)

        warnings_list = []
        records = []
        seen_emails = set()
        missing_email_count = 0
        duplicate_count = 0
        for _, row in df.iterrows():
            name  = row.get(name_col)
            email = row.get(email_col)
            if not isinstance(name, str) or not name.strip():
                continue
            name = smart_title(name.strip())
            # Resolve email; treat NaN/None as missing
            if email is None or (isinstance(email, float) and pd.isna(email)):
                email_str = ""
            else:
                email_str = str(email).strip().lower()
            email_valid = bool(email_str) and looks_like_email(email_str)
            if not email_valid:
                missing_email_count += 1
            if email_valid and email_str in seen_emails:
                duplicate_count += 1
            elif email_valid:
                seen_emails.add(email_str)
            for col in session_cols:
                if _is_yes_like(row.get(col)):
                    records.append({
                        "NAME":          name,
                        "EMAIL":         email_str,
                        "SESSION_LABEL": str(col).strip(),
                        "ATTENDED":      True,
                    })

        rows = pd.DataFrame(records,
                            columns=["NAME","EMAIL","SESSION_LABEL","ATTENDED"])

        if missing_email_count:
            warnings_list.append(
                f"{missing_email_count} row(s) had no valid email address.")
        if duplicate_count:
            warnings_list.append(
                f"{duplicate_count} duplicate email(s) found in roster.")
        ignored = [str(c) for c in df.columns
                   if c not in (name_col, email_col) and c not in session_cols
                   and not _looks_like_unnamed(c)]
        if ignored:
            warnings_list.append(
                "Ignored columns (not used for certificates): "
                + ", ".join(ignored))

        return CanonicalAttendance(
            rows=rows,
            granularity="per_participant",
            training_title=_title_from_filename(path),
            session_columns=[(str(c).strip(), None, None) for c in session_cols],
            source_path=path,
            profile_used=self.name,
            confidence=self.detect(path),
            warnings=warnings_list,
        )


# Order matters only as a tiebreaker — auto_detect_profile returns the
# highest-scoring match, but if two profiles tie, the earlier one wins.
DEFAULT_PROFILES = [
    QualtricsLcSignoutProfile(),
    QualtricsLcSigninProfile(),
    QualtricsStandardProfile(),
    PerParticipantRosterProfile(),
]


def auto_detect_profile(path: str, profiles=None):
    """Return (profile, confidence). profile is None if no profile claims
    the file with confidence > 0."""
    if profiles is None:
        profiles = DEFAULT_PROFILES
    best = (None, 0.0)
    for p in profiles:
        try:
            c = float(p.detect(path))
        except Exception:
            c = 0.0
        if c > best[1]:
            best = (p, c)
    return best


def ingest_attendance(path: str, profile: Profile = None) -> CanonicalAttendance:
    """Parse `path` into a CanonicalAttendance.

    If `profile` is None, auto-detect the best-matching profile from
    DEFAULT_PROFILES. Raises ValueError when no profile claims the file —
    in Phase 3 this is where the manual-mapping fallback will hook in."""
    if profile is None:
        profile, _ = auto_detect_profile(path)
    if profile is None:
        raise ValueError(
            f"No ingest profile matched {os.path.basename(path)}. "
            "Manual column mapping will be added in a future phase.")
    return profile.parse(path)


# ---------- DOCX certificate generation ----------

def generate_one(name: str, email: str, event: dict, template: str, outdir: str, ceu_vars=None):
    os.makedirs(outdir, exist_ok=True)
    mapping = {
        "<<NAME>>": name,
        "<<TRAINING>>": event["TRAINING"],
        "<<DATE>>": event["DATE"],
        "<<TRAINER>>": event["TRAINER"],
        "<<HOURS>>": str(event["HOURS"]),
        "<<LOCATION>>": event["LOCATION"],
    }
    if ceu_vars:
        mapping.update(ceu_vars)
    doc = Document(template)
    replace_placeholders_everywhere(doc, mapping)
    safe_name = re.sub(r"[^A-Za-z0-9._ -]", "", name)
    safe_training = safe_filename(event['TRAINING'])
    base = f"{safe_name} - {safe_training}".strip().replace("  ", " ")
    out_docx = os.path.join(outdir, base + ".docx")
    doc.save(out_docx)
    out_file = out_docx
    if docx2pdf_convert is not None:
        try:
            out_pdf = os.path.join(outdir, base + ".pdf")
            docx2pdf_convert(out_docx, out_pdf)
            if os.path.isfile(out_pdf): out_file = out_pdf
        except Exception:
            pass
    return out_file

def match_participants(si: pd.DataFrame, so: pd.DataFrame, method: int):
    if method == 1:
        si2 = si.copy() if si is not None else pd.DataFrame(columns=["NAME","EMAIL"])
        so2 = so.copy() if so is not None else pd.DataFrame(columns=["NAME","EMAIL"])
        if si2.empty or so2.empty:
            final = pd.DataFrame(columns=["NAME","EMAIL"])
        else:
            si2["KEY"] = si2["EMAIL"].fillna("").str.lower() + si2["NAME"].str.lower()
            so2["KEY"] = so2["EMAIL"].fillna("").str.lower() + so2["NAME"].str.lower()
            keys = set(si2["KEY"]) & set(so2["KEY"])
            final = si2[si2["KEY"].isin(keys)][["NAME", "EMAIL"]].copy()
    elif method == 2:
        final = si.copy() if si is not None else pd.DataFrame(columns=["NAME","EMAIL"])
    else:
        final = so.copy() if so is not None else pd.DataFrame(columns=["NAME","EMAIL"])
    if not final.empty:
        final = final.drop_duplicates(subset=["EMAIL"], keep="first")
        final = final.drop_duplicates(subset=["NAME"], keep="first").reset_index(drop=True)
    return final

# ---------- Training Request helpers (standalone) ----------

def _req_save_json(store_dir: str, data: dict):
    """Save/overwrite a request JSON file keyed by data['id']."""
    os.makedirs(store_dir, exist_ok=True)
    fpath = os.path.join(store_dir, f"{data['id']}.json")
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _req_load_json(store_dir: str, req_id: str) -> dict:
    """Load a single request by ID. Returns None if not found."""
    fpath = os.path.join(store_dir, f"{req_id}.json")
    if not os.path.isfile(fpath):
        return None
    with open(fpath, encoding="utf-8") as f:
        return json.load(f)


def _req_load_all_json(store_dir: str) -> list:
    """Return all requests sorted by submitted_at (newest first)."""
    if not os.path.isdir(store_dir):
        return []
    records = []
    for fname in os.listdir(store_dir):
        if fname.endswith(".json"):
            try:
                with open(os.path.join(store_dir, fname), encoding="utf-8") as f:
                    records.append(json.load(f))
            except Exception:
                pass
    records.sort(key=lambda r: r.get("submitted_at", ""), reverse=True)
    return records


def _req_append_to_master(mtl_path: str, data: dict):
    """Append a new row to the Open Training Requests tab."""
    if load_workbook is None:
        return
    wb = load_workbook(mtl_path)
    ws = wb["Open Training Requests"]
    # Columns: Agency, Agency Contact, Contact Information,
    #          Training Requested, Quote, Date Quote Emailed, Staff, Notes/Updates
    notes = data.get("additional_comments", "")
    row_vals = [
        data.get("org_name", ""),
        data.get("billing_contact", ""),
        data.get("billing_email", "") or data.get("billing_phone", ""),
        data.get("training_requested", ""),
        "",                                         # Quote (blank until Dr. Sprang quotes)
        "",                                         # Date Quote Emailed
        data.get("preferred_trainer", ""),
        f"Requested: {data.get('requested_date','')}  |  "
        f"Location: {data.get('location','')}  |  "
        f"Participants: {data.get('num_participants','')}  |  "
        f"ID: {data.get('id','')}  |  {notes}".strip(" |"),
    ]
    ws.append(row_vals)
    wb.save(mtl_path)


def _req_update_master_quote(mtl_path: str, org_name: str,
                              quote_total: str, quote_date: str):
    """Update Quote + Date Quote Emailed columns for the matching org row."""
    if load_workbook is None:
        return
    wb = load_workbook(mtl_path)
    ws = wb["Open Training Requests"]
    for row in ws.iter_rows(min_row=3):
        cell_agency = row[0].value
        if cell_agency and str(cell_agency).strip().lower() == org_name.strip().lower():
            row[4].value = f"${quote_total}"    # Quote column
            row[5].value = quote_date           # Date Quote Emailed
            break
    wb.save(mtl_path)


def _req_resolve_in_master(mtl_path: str, data: dict, outcome: str):
    """Move the row from Open to Resolved Training Requests."""
    if load_workbook is None:
        return
    wb = load_workbook(mtl_path)
    ws_open = wb["Open Training Requests"]
    ws_res  = wb["Resolved Training Requests"]
    org_name = data.get("org_name", "").strip().lower()

    row_to_move = None
    row_idx = None
    for i, row in enumerate(ws_open.iter_rows(min_row=3), start=3):
        if row[0].value and str(row[0].value).strip().lower() == org_name:
            row_to_move = [c.value for c in row]
            row_idx = i
            break

    if row_to_move is None:
        # Nothing found in Open — just append to Resolved directly
        row_to_move = [
            data.get("org_name", ""),
            data.get("billing_contact", ""),
            data.get("billing_email", ""),
            data.get("training_requested", ""),
            data.get("quote_total", ""),
            data.get("quoted_at", "")[:10] if data.get("quoted_at") else "",
        ]

    # Append to Resolved
    # Columns: Agency, Agency Contact, Contact Information, Training Requested,
    #          Quote, Date Quote Emailed, Expires, Trainer Scheduled, Outcome
    res_row = list(row_to_move[:6]) + [
        "",
        data.get("preferred_trainer", ""),
        outcome.capitalize(),
    ]
    ws_res.append(res_row)

    # Delete from Open (if found)
    if row_idx is not None:
        ws_open.delete_rows(row_idx)

    wb.save(mtl_path)


def generate_request_pdf(data: dict, output_path: str) -> str:
    """Generate the training request form as a PDF using reportlab."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=1.0 * inch, rightMargin=1.0 * inch,
        topMargin=0.9 * inch, bottomMargin=0.9 * inch,
    )
    styles = getSampleStyleSheet()

    title_sty = ParagraphStyle("req_title", fontSize=16, fontName="Helvetica-Bold",
                                alignment=TA_CENTER, spaceAfter=4)
    org_sty   = ParagraphStyle("req_org",   fontSize=10, fontName="Helvetica-Bold",
                                alignment=TA_CENTER, spaceAfter=14)
    lbl_sty   = ParagraphStyle("req_lbl",   fontSize=9, fontName="Helvetica-Bold",
                                textColor=colors.HexColor("#003366"))
    val_sty   = ParagraphStyle("req_val",   fontSize=10, fontName="Helvetica",
                                leftIndent=12, spaceAfter=6)
    sec_sty   = ParagraphStyle("req_sec",   fontSize=11, fontName="Helvetica-Bold",
                                textColor=colors.HexColor("#003366"),
                                spaceBefore=10, spaceAfter=4)

    story = []

    # Header
    story.append(Paragraph("Training Request Form", title_sty))
    story.append(Paragraph("University of Kentucky Center on Trauma and Children", org_sty))
    story.append(HRFlowable(width="100%", thickness=1.5,
                             color=colors.HexColor("#003366"), spaceAfter=8))

    def field(label, value):
        story.append(Paragraph(label, lbl_sty))
        story.append(Paragraph(value or "—", val_sty))

    field("Today's Date:", data.get("request_date", ""))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey, spaceAfter=6))

    field("Who (Organization):", data.get("org_name", ""))
    field("What (Training Requested):", data.get("training_requested", ""))
    field("When (Requested Date):", data.get("requested_date", ""))
    field("Where (Location):", data.get("location", ""))

    story.append(Paragraph("Participant Information", sec_sty))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey, spaceAfter=4))
    field("# of Participants:", data.get("num_participants", ""))
    field("Roles of Participants:", data.get("participant_roles", ""))

    story.append(Paragraph("Logistics", sec_sty))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey, spaceAfter=4))
    field("Potential Trainer:", data.get("preferred_trainer", ""))
    field("Mileage Cost:", data.get("mileage_cost", ""))
    field("Additional Travel Cost (Hotel, Parking, etc.):", data.get("other_travel_cost", ""))

    story.append(Paragraph("Billing Contact Information", sec_sty))
    story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey, spaceAfter=4))
    field("Contact Name:", data.get("billing_contact", ""))
    field("Contact Email:", data.get("billing_email", ""))
    field("Contact Phone:", data.get("billing_phone", ""))

    comments = data.get("additional_comments", "").strip()
    if comments:
        story.append(Paragraph("Additional Comments", sec_sty))
        story.append(HRFlowable(width="100%", thickness=0.4, color=colors.grey, spaceAfter=4))
        safe_comments = (comments.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"))
        story.append(Paragraph(safe_comments, val_sty))

    # Quote placeholder
    story.append(Spacer(1, 0.2 * inch))
    story.append(HRFlowable(width="100%", thickness=1.5,
                             color=colors.HexColor("#003366"), spaceAfter=8))
    story.append(Paragraph("Quote:  $___________________", sec_sty))
    story.append(Paragraph("Additional Comments/Instructions:", sec_sty))
    story.append(Spacer(1, 0.5 * inch))

    doc.build(story)
    return output_path


def generate_quote_pdf(data: dict, output_path: str) -> str:
    """Generate the training quote as a PDF using reportlab (mirrors CTAC quote template)."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        leftMargin=1.0 * inch, rightMargin=1.0 * inch,
        topMargin=0.9 * inch, bottomMargin=0.9 * inch,
    )

    hdr_sty   = ParagraphStyle("q_hdr",  fontSize=10, fontName="Helvetica",
                                alignment=TA_RIGHT, leading=14)
    title_sty = ParagraphStyle("q_title", fontSize=18, fontName="Helvetica-Bold",
                                alignment=TA_CENTER, spaceAfter=4)
    prep_sty  = ParagraphStyle("q_prep",  fontSize=11, fontName="Helvetica",
                                spaceAfter=16)
    footer_sty= ParagraphStyle("q_foot",  fontSize=9,  fontName="Helvetica-Oblique",
                                textColor=colors.grey, spaceBefore=12)

    story = []

    # Org header (right-aligned block like the template)
    hdr_text = (
        "<b>UK – Center on Trauma and Children</b><br/>"
        "3470 Blazer Parkway, Suite 100<br/>"
        "Lexington, KY 40509<br/>"
        "859-218-6901"
    )
    story.append(Paragraph(hdr_text, hdr_sty))
    story.append(Spacer(1, 0.08 * inch))

    # Date
    today_str = datetime.now().strftime("%m/%d/%Y")
    story.append(Paragraph(f"Date:  {today_str}", prep_sty))

    # Title
    story.append(Paragraph("Training Quote", title_sty))
    story.append(HRFlowable(width="100%", thickness=1.5,
                             color=colors.HexColor("#003366"), spaceAfter=10))

    # Prepared for
    bill_name = data.get("billing_contact", "") or data.get("org_name", "")
    org_name  = data.get("org_name", "")
    prep_for  = bill_name
    if org_name and org_name != bill_name:
        prep_for = f"{bill_name} – {org_name}"
    story.append(Paragraph(f"Prepared for:  {prep_for}", prep_sty))

    # Line items table
    tbl_data = [["Training Event(s)", "Date", "Cost"]]
    line_items = data.get("line_items", [])
    total = 0.0
    for li in line_items:
        cost_str = li.get("cost", "").replace("$","").replace(",","").strip()
        cost_val = 0.0
        try:
            cost_val = float(cost_str or 0)
        except ValueError:
            pass
        total += cost_val
        tbl_data.append([
            li.get("desc", ""),
            li.get("date", ""),
            f"${cost_val:,.2f}" if cost_val else "",
        ])
    # Pad to at least 4 rows
    while len(tbl_data) < 5:
        tbl_data.append(["", "", ""])
    # Total row
    tbl_data.append(["", "Total", f"${total:,.2f}"])

    col_w = [4.0 * inch, 1.4 * inch, 1.1 * inch]
    tbl = Table(tbl_data, colWidths=col_w)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1,  0), colors.HexColor("#003366")),
        ("TEXTCOLOR",     (0, 0), (-1,  0), colors.white),
        ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 0), (-1, -1), 10),
        ("ALIGN",         (1, 0), (-1, -1), "CENTER"),
        ("ALIGN",         (0, 0), ( 0, -1), "LEFT"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2),
         [colors.white, colors.HexColor("#f0f4fa")]),
        ("BACKGROUND",    (0, -1),(-1, -1), colors.HexColor("#e8eef7")),
        ("FONTNAME",      (1, -1),(-1, -1), "Helvetica-Bold"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), ( 0, -1), 8),
    ]))
    story.append(tbl)

    # Footer note
    story.append(Paragraph(
        "Questions? Please email joshua.fisherkeller@uky.edu and trish.polly@uky.edu",
        footer_sty
    ))

    doc.build(story)
    return output_path


# ---------- Evaluation report helpers ----------

def parse_eval_file(path: str):
    """Parse a Qualtrics evaluation CSV or XLSX.
    Returns ({}, sessions_dict) where sessions_dict maps date_str -> DataFrame.
    The first data row is skipped (Qualtrics label row).
    """
    if path.lower().endswith((".xlsx", ".xls")):
        raw = pd.read_excel(path, header=0, dtype=str)
    else:
        raw = pd.read_csv(path, header=0, dtype=str)
    if len(raw) < 2:
        raise ValueError("Evaluation file appears empty or has no data rows.")
    # Row 0 is the Qualtrics question-label row; actual data starts at row 1
    data = raw.iloc[1:].copy()
    data.columns = raw.columns
    if "RecordedDate" not in data.columns:
        raise ValueError("No 'RecordedDate' column found.  Is this a Qualtrics export?")
    data["_date"] = pd.to_datetime(data["RecordedDate"], errors="coerce").dt.date
    data = data[data["_date"].notna()].copy()
    sessions = {}
    for date, grp in data.groupby("_date"):
        sessions[str(date)] = grp.reset_index(drop=True)
    return {}, sessions


_EVAL_JUNK = {
    "na", "n/a", "n-a", "n/a.", "none", "nothing", "no", "n", ".", "-",
    "--", "...", "er", "wer", "qwe", "jfkf", "jdjf", "udud", "ddd", "test",
    "nope", "idk", "b/a", "not applicable", "not applicable.", "na.",
}


def _eval_is_meaningful(s) -> bool:
    if not isinstance(s, str):
        return False
    s = s.strip()
    if len(s) < 5:
        return False
    return s.lower() not in _EVAL_JUNK


_EVAL_RATING_COLS = ["Q51_1", "Q51_2", "Q51_3", "Q51_4", "Q51_5", "Q51_6"]
_EVAL_RATING_LABELS = [
    "Trainer was effective",
    "High level of consistency between content and objectives",
    "I will be able to incorporate the knowledge and skills I have gained from this training into my daily work",
    "I am satisfied with the level of practical knowledge and skills presented in this training",
    "Teaching methods appropriate for intended audience",
    "Teaching methods appropriate for subject matter",
]


def generate_eval_pdf(training_name: str, sessions_data: list,
                      output_path: str, notes: str = "") -> str:
    """
    Generate a training evaluation report PDF.
    sessions_data: list of (session_label: str, df: pd.DataFrame)
    Returns output_path.
    """
    doc = SimpleDocTemplate(
        output_path, pagesize=letter,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
        topMargin=0.75 * inch, bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    title_sty = ParagraphStyle("EvalTitle", parent=styles["Title"],
                                fontSize=16, spaceAfter=4, alignment=TA_CENTER)
    org_sty   = ParagraphStyle("OrgName",   parent=styles["Normal"],
                                fontSize=10, alignment=TA_CENTER, spaceAfter=2,
                                textColor=colors.grey)
    sess_sty  = ParagraphStyle("SessHdr",   parent=styles["Heading2"],
                                fontSize=13, spaceBefore=14, spaceAfter=6,
                                textColor=colors.HexColor("#003366"))
    q_sty     = ParagraphStyle("QHdr",      parent=styles["Heading3"],
                                fontSize=11, spaceBefore=10, spaceAfter=3)
    cnt_sty   = ParagraphStyle("Cnt",       parent=styles["Normal"],
                                fontSize=9,  spaceAfter=4, textColor=colors.grey)
    resp_sty  = ParagraphStyle("Resp",      parent=styles["Normal"],
                                fontSize=10, leftIndent=14, spaceAfter=2, leading=14)
    note_sty  = ParagraphStyle("Note",      parent=styles["Normal"],
                                fontSize=10, spaceAfter=4)

    story = []

    # Header
    story.append(Paragraph(training_name or "Training Evaluation Report", title_sty))
    story.append(Paragraph("University of Kentucky Center on Trauma and Children", org_sty))
    story.append(Spacer(1, 0.1 * inch))
    story.append(HRFlowable(width="100%", thickness=1.5,
                             color=colors.HexColor("#003366")))
    story.append(Spacer(1, 0.1 * inch))

    if notes.strip():
        story.append(Paragraph("<b>Coordinator Notes:</b>", q_sty))
        for line in notes.strip().split("\n"):
            if line.strip():
                safe_line = (line.strip()
                             .replace("&", "&amp;")
                             .replace("<", "&lt;")
                             .replace(">", "&gt;"))
                story.append(Paragraph(safe_line, note_sty))
        story.append(Spacer(1, 0.1 * inch))

    for session_label, df in sessions_data:
        story.append(Paragraph(session_label, sess_sty))

        # Ratings table
        rating_cols_present = [c for c in _EVAL_RATING_COLS if c in df.columns]
        if rating_cols_present:
            rating_num = df[rating_cols_present].apply(pd.to_numeric, errors="coerce")
            n_responses = int(rating_num.dropna(how="all").shape[0])
            story.append(Paragraph("Q51 - Please rate", q_sty))
            story.append(Paragraph(f"{n_responses} Responses", cnt_sty))

            tbl_data = [["Field", "Min", "Max", "Mean"]]
            for col, label in zip(_EVAL_RATING_COLS, _EVAL_RATING_LABELS):
                if col not in df.columns:
                    continue
                vals = pd.to_numeric(df[col], errors="coerce").dropna()
                if vals.empty:
                    continue
                display = label if len(label) <= 72 else label[:70] + "\u2026"
                tbl_data.append([display,
                                  f"{vals.min():.2f}",
                                  f"{vals.max():.2f}",
                                  f"{vals.mean():.2f}"])
            if len(tbl_data) > 1:
                col_widths = [3.9 * inch, 0.65 * inch, 0.65 * inch, 0.65 * inch]
                t = Table(tbl_data, colWidths=col_widths)
                t.setStyle(TableStyle([
                    ("BACKGROUND",     (0, 0), (-1,  0), colors.HexColor("#003366")),
                    ("TEXTCOLOR",      (0, 0), (-1,  0), colors.white),
                    ("FONTNAME",       (0, 0), (-1,  0), "Helvetica-Bold"),
                    ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE",       (0, 0), (-1, -1), 9),
                    ("ALIGN",          (1, 0), (-1, -1), "CENTER"),
                    ("ALIGN",          (0, 0), ( 0, -1), "LEFT"),
                    ("GRID",           (0, 0), (-1, -1), 0.4, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.white, colors.HexColor("#f5f5f5")]),
                    ("TOPPADDING",     (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
                    ("LEFTPADDING",    (0, 0), ( 0, -1), 6),
                ]))
                story.append(t)
                story.append(Spacer(1, 0.08 * inch))

        # Open-ended questions
        for col, label in [
            ("Q51", "Q51 - What part of the training was the most helpful?"),
            ("Q52", "Q52 - What are changes you would make to improve this training?"),
            ("Q53", "Q53 - Additional Comments"),
        ]:
            if col not in df.columns:
                continue
            responses = [str(v).strip() for v in df[col] if _eval_is_meaningful(str(v))]
            if not responses:
                continue
            story.append(Paragraph(label, q_sty))
            story.append(Paragraph(f"{len(responses)} Responses", cnt_sty))
            for resp in responses:
                safe = (resp.replace("&", "&amp;")
                            .replace("<", "&lt;")
                            .replace(">", "&gt;"))
                story.append(Paragraph(safe, resp_sty))
            story.append(Spacer(1, 0.06 * inch))

    doc.build(story)
    return output_path


# ---------- main ----------

if __name__ == "__main__":
    import traceback as _tb
    _log = _os.path.join(_os.path.expanduser("~"), "Documents", "TEM_startup_error.log")
    try:
        app = App()
        app.mainloop()
    except Exception:
        err = _tb.format_exc()
        try:
            with open(_log, "w", encoding="utf-8") as _f:
                _f.write(err)
        except Exception:
            pass
        # Try to show a messagebox with the error
        try:
            import tkinter as _tk2, tkinter.messagebox as _mb2
            _r = _tk2.Tk(); _r.withdraw()
            _mb2.showerror("Startup Error",
                f"The program crashed on startup.\n\nError:\n{err[:1500]}\n\n"
                f"Details saved to:\n{_log}")
            _r.destroy()
        except Exception:
            pass
